# -*- coding: utf-8 -*-
"""해상 정시성(Global Schedule Reliability) '다음 달' 예측 — 순수 추가 모듈(add-only).

설계: 숫자는 코드가 계산(추세 + 모멘텀 + 계절성), 해석 문장만 Anthropic API가 작성.
계절성이 뚜렷한 지표라 계절성 가중치를 가장 높게 둔다. 백분율이라 0~100%로 clamp.

★ 스폰지 예측(icis_forecast.py)과 공통화하지 않고 별도로 작성(중복 허용).
★ 기존 코드/데이터(update_schedule_reliability, 차트 등)는 전혀 건드리지 않는다.
"""
import os
import re
import json
import statistics
import datetime

import requests

SRF_XLSX_URL = ("https://www.sea-intelligence.com/images/press_docs/GLP-May2026/"
                "20260527_-_Sea-Intelligence_GLP_Press_Release_-_May_2026.xlsx")
SRF_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _srf_fetch():
    """Sea-Intelligence xlsx 'Fig 1' 파싱 → years {year:[12 %|None]}. 실패 시 None.
       (기존 update_schedule_reliability 와 동일 로직을 여기 복제 — 공유하지 않음)."""
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"}
    try:
        import io
        import openpyxl
        # 순수 추가: 최신 GLP 편 URL 을 자동 탐색해 쓴다(캐시됨). 실패 시 아래 고정 URL.
        # ★ 예측 계산 로직은 그대로다 — 입력 데이터만 최신으로 바뀐다.
        url = SRF_XLSX_URL
        try:
            from sea_intelligence import si_latest_xlsx_url
            url = si_latest_xlsx_url() or SRF_XLSX_URL
        except Exception:  # noqa: BLE001
            pass
        r = requests.get(url, headers=hdr, allow_redirects=True, timeout=15)
        r.raise_for_status()
        if r.content[:2] != b"PK":
            return None
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        if "Fig 1" not in wb.sheetnames:
            return None
        ws = wb["Fig 1"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        month_cols = [i for i, v in enumerate(header)
                      if isinstance(v, str) and v.strip()[:3] in SRF_MONTHS]
        if len(month_cols) < 12:
            month_cols = list(range(1, 13))
        years = {}
        for row in ws.iter_rows(min_row=2):
            cells = [c.value for c in row]
            if not cells:
                continue
            try:
                yi = int(cells[0])
            except (TypeError, ValueError):
                continue
            if not (2000 <= yi <= 2100):
                continue
            vals = []
            for ci in month_cols:
                v = cells[ci] if ci < len(cells) else None
                if isinstance(v, (int, float)):
                    vals.append(round(v * 100.0 if v <= 1.5 else float(v), 1))
                else:
                    vals.append(None)
            years[str(yi)] = vals
        return years or None
    except Exception:  # noqa: BLE001
        return None


def _linreg_next(points, x_next):
    """(x, y) 점들로 단순선형회귀 → x_next 예측값."""
    n = len(points)
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    xb, yb = sum(xs) / n, sum(ys) / n
    den = sum((x - xb) ** 2 for x in xs)
    if den == 0:
        return yb
    slope = sum((xs[i] - xb) * (ys[i] - yb) for i in range(n)) / den
    return slope * x_next + (yb - slope * xb)


def _srf_load_key():
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key
    here = os.path.dirname(os.path.abspath(__file__))
    for path in (os.path.join(here, ".env"), ".env"):
        try:
            with open(path, encoding="utf-8") as f:
                for line in f:
                    s = line.strip()
                    if s.startswith("ANTHROPIC_API_KEY"):
                        return s.split("=", 1)[1].strip().strip('"').strip("'")
        except OSError:
            pass
    return None


def _srf_ai(target_month, recent12, samemonth_hist, path, months_ahead, o):
    """Anthropic API로 해석 문장만 생성. 실패 시 None(→ 통계값만 표시)."""
    key = _srf_load_key()
    if not key:
        print("[sr_forecast] ANTHROPIC_API_KEY 없음 → AI 해석 생략(통계 예측값만 표시)")
        return None
    hist = ", ".join("%d년 %.1f%%" % (y, v) for y, v in samemonth_hist) or "없음"
    pathtxt = " → ".join("%s %.1f" % (p["label"], p["v"]) for p in path)
    prompt = (
        "너는 글로벌 해상 정시성(정시 도착 비율) 애널리스트다. 최신 실측은 %s(직전 실측), "
        "예측 대상은 %s로 %d개월 후를 1개월씩 이어붙이는 다단계(재귀) 방식으로 추정했다. "
        "이 수치를 해석하는 짧은 문장만 작성하라.\n\n"
        "최근 12개월 실제 정시성(%%): %s\n"
        "대상월(%s)의 과거 연도 동월 실측: %s\n"
        "다단계 경로(추정): %s\n"
        "최종 예측=%.1f%% (신뢰구간 %.1f~%.1f), 직전 실측 대비=%+.1f%%p, 전년 동월 대비=%s\n\n"
        "제약:\n"
        "- 제시된 수치 외의 숫자를 만들어내지 말 것.\n"
        "- 특정 선사·항만·파업·지정학 이벤트를 아는 척하지 말 것(학습 시점이 오래됐고 해운은 외부 충격에 크게 좌우됨).\n"
        "- %d개월 앞 예측이라 계절 패턴 의존도가 높고 불확실성이 크다는 점을 반영할 것.\n"
        "- 단정하지 말고 '~할 전망' 같은 추정 어조.\n"
        "- 한국어.\n\n"
        "아래 JSON만 출력(코드펜스·설명 없이 JSON only):\n"
        '{"summary":"현재 정시성 수준과 대상월 전망 2~3문장",'
        '"seasonal":"계절 패턴상 이 시기의 일반적 흐름 1~2문장",'
        '"yoy":"전년 동월 대비 평가 1문장","risk":"상방|하방|중립","caution":"예측의 한계 1문장"}'
    ) % (o["last_data_month"], target_month, months_ahead, recent12, target_month, hist, pathtxt,
         o["predict"], o["ci_low"], o["ci_high"], o["delta_pp"],
         ("%+.1f%%p" % o["yoy_pp"] if o["yoy_pp"] is not None else "데이터 없음"), months_ahead)
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": key, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": "claude-sonnet-4-6", "max_tokens": 1024,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=30)
        r.raise_for_status()
        text = "".join(b.get("text", "") for b in r.json().get("content", [])
                       if b.get("type") == "text").strip()
        print("[sr_forecast] AI 응답 원문:\n%s" % text)
        cleaned = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.I | re.M).strip()
        data = json.loads(cleaned)
        print("[sr_forecast] AI 파싱 성공")
        return data
    except Exception as e:  # noqa: BLE001
        print("[sr_forecast] AI 실패(%s) → 통계 예측값만 표시" % e)
        return None


def _srf_seasonal_deltas(years, from_idx, to_idx):
    """과거 연도들의 (from_idx월 → to_idx월) 변화폭(%p) 목록. 연말→연초(Dec→Jan) 처리 포함."""
    out = []
    if from_idx != 11 or to_idx != 0:  # 같은 해 인접 월
        for y in years:
            a, b = years[y][from_idx], years[y][to_idx]
            if a is not None and b is not None:
                out.append(b - a)
    else:  # Dec(해 Y) → Jan(해 Y+1)
        for y in years:
            ny = str(int(y) + 1)
            if ny in years and years[y][11] is not None and years[ny][0] is not None:
                out.append(years[ny][0] - years[y][11])
    return out


def compute_sr_forecast():
    """FETCHERS용 진입점. '현재 시스템 날짜의 다음 달'까지 다단계(재귀) 예측 + AI 해석.
       마지막 데이터 월 → 대상 월 간격을 코드가 계산하고, 1개월 앞 예측을 반복해 이어붙인다.
       단계가 진행될수록 계절성 가중치↑(추세·모멘텀↓), 신뢰구간은 √단계수로 확대. 0~100% clamp."""
    years = _srf_fetch()
    if not years:
        print("[sr_forecast] 정시성 데이터 로드 실패 → 예측 섹션 표시 안 함")
        return {"status": "error", "reason": "정시성 데이터 없음"}

    yrs = sorted(years.keys())
    flat = [(int(y), m, years[y][m]) for y in yrs for m in range(12)]  # (year, monthIdx, value)
    last_pos = next((i for i in range(len(flat) - 1, -1, -1) if flat[i][2] is not None), None)
    if last_pos is None:
        return {"status": "error", "reason": "유효 정시성 값 없음"}
    ly, lm, lv = flat[last_pos]
    last_data_month = "%04d-%02d" % (ly, lm + 1)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 예측 대상 = 현재 시스템 날짜의 '다음 달'(하드코딩 금지, 실행 시점 기준 자동)
    today = datetime.date.today()
    ty, tmidx = (today.year + 1, 0) if today.month == 12 else (today.year, today.month)  # 다음 달 index
    target_month = "%04d-%02d" % (ty, tmidx + 1)
    steps = (ty - ly) * 12 + (tmidx - lm)  # 마지막 데이터 → 대상 월 간격(개월)
    if steps < 1:  # 데이터가 이미 대상월 이상이면 최소 1개월 앞
        steps = 1
        ty, tmidx = (ly + 1, 0) if lm == 11 else (ly, lm + 1)
        target_month = "%04d-%02d" % (ty, tmidx + 1)
    print("[sr_forecast] 최신 데이터 %s(%.1f%%), 현재 %s → 대상 월 %s (%d개월 후, 다단계 %d회)"
          % (last_data_month, lv, today.strftime("%Y-%m"), target_month, steps, steps))

    # 결측: 최근 3개월 유효값 2개 미만이면 예측 생략
    recent3 = [flat[i][2] for i in range(max(0, last_pos - 2), last_pos + 1)]
    if len([v for v in recent3 if v is not None]) < 2:
        print("[sr_forecast] 최근 3개월 유효값 부족 → 데이터 부족")
        return {"status": "ok", "forecast_status": "insufficient",
                "reason": "최근 3개월 유효값 부족", "target_month": target_month,
                "last_data_month": last_data_month, "months_ahead": steps, "generated_at": now}

    # ── 재귀 다단계 예측: 4월 실측 → 5월 → 6월 → 7월 → 8월 ──
    chron = [v for (_, _, v) in flat[:last_pos + 1]]  # 실측(Nones 포함)
    cur_y, cur_m, cur_v = ly, lm, lv
    path, sigma_base = [], None
    for s in range(1, steps + 1):
        nmi = (cur_m + 1) % 12
        nyi = cur_y + (1 if cur_m == 11 else 0)
        # a) 추세: chron 최근 6개 유효점 선형회귀
        last6 = chron[-6:]
        pts = [(k, last6[k]) for k in range(len(last6)) if last6[k] is not None]
        a = _linreg_next(pts, len(last6)) if len(pts) >= 2 else cur_v
        # b) 모멘텀: 최근 3개 변화폭 평균
        vseq = [v for v in chron if v is not None]
        diffs = [vseq[i] - vseq[i - 1] for i in range(1, len(vseq))]
        b = cur_v + (sum(diffs[-3:]) / len(diffs[-3:]) if diffs else 0.0)
        # c) 계절성: 과거 연도 (cur_m월 → nmi월) 변화폭 평균을 현재값에 적용
        sd = _srf_seasonal_deltas(years, cur_m, nmi)
        seas = sum(sd) / len(sd) if sd else 0.0
        c = cur_v + seas
        sig = statistics.pstdev(sd) if len(sd) >= 2 else 0.0
        if sigma_base is None:
            sigma_base = sig  # 1단계 계절 변화폭 표준편차 = 기본 σ
        # 가중치: 단계 진행 → 계절성↑ / 추세·모멘텀↓ (5월0.4 6월0.6 7월0.8 8월1.0)
        sw = min(1.0, 0.4 + 0.2 * (s - 1))
        tw = mw = (1.0 - sw) / 2.0
        pred = max(0.0, min(100.0, tw * a + mw * b + sw * c))
        cum_ci = (sigma_base or 0.0) * (s ** 0.5)  # 누적 구간폭 = 기본σ×√단계
        print("[sr_forecast]  %d단계 %04d-%02d | 입력(직전값)=%.1f | a=%.1f b=%.1f c=%.1f(계절Δ%+.2f n=%d) | 가중 추세%.1f·모멘텀%.1f·계절%.1f → %.1f%% | 누적CI ±%.2f(%.1f~%.1f)"
              % (s, nyi, nmi + 1, cur_v, a, b, c, seas, len(sd), tw, mw, sw, pred, cum_ci,
                 max(0.0, pred - cum_ci), min(100.0, pred + cum_ci)))
        path.append({"m": "%04d-%02d" % (nyi, nmi + 1), "label": "%d월" % (nmi + 1), "v": round(pred, 1)})
        chron.append(pred)
        cur_y, cur_m, cur_v = nyi, nmi, pred

    final = round(cur_v, 1)
    total_sigma = (sigma_base or 0.0) * (steps ** 0.5)  # 기본 σ × √(단계 수)
    ci_low = round(max(0.0, cur_v - total_sigma), 1)
    ci_high = round(min(100.0, cur_v + total_sigma), 1)  # 0~100 clamp 유지
    delta_pp = round(final - lv, 1)  # 최신 실측(마지막 데이터월) 대비
    yoy_prev = years[str(ty - 1)][tmidx] if str(ty - 1) in years else None
    yoy_pp = round(final - yoy_prev, 1) if yoy_prev is not None else None
    risk = "상방" if delta_pp > 1 else ("하방" if delta_pp < -1 else "중립")

    # 검증: 과거 동월(대상월) 실측과 비슷한 범위인지
    tgt_hist = [(int(y), years[y][tmidx]) for y in yrs if years[y][tmidx] is not None and int(y) < ty]
    if tgt_hist:
        lo = min(v for _, v in tgt_hist)
        hi = max(v for _, v in tgt_hist)
        ok = (lo - total_sigma) <= final <= (hi + total_sigma)
        print("[sr_forecast] 과거 %s 실측: %s | 예측 %.1f%% (과거 범위 %.1f~%.1f) → %s"
              % (SRF_MONTHS[tmidx], ", ".join("%d년 %.1f" % (y, v) for y, v in tgt_hist),
                 final, lo, hi, "타당(범위 내/근접)" if ok else "범위 밖(주의)"))
    print("[sr_forecast] 최종 %s = %.1f%% | 기본σ=%.2f × √%d = ±%.2f%%p | CI %.1f~%.1f | 직전 실측(%s %.1f) 대비 %+.1f%%p"
          % (target_month, final, sigma_base or 0.0, steps, total_sigma, ci_low, ci_high,
             last_data_month, lv, delta_pp))
    if yoy_pp is not None:
        print("[sr_forecast] 전년 동월(%s %d) %.1f%% 대비 %+.1f%%p" % (SRF_MONTHS[tmidx], ty - 1, yoy_prev, yoy_pp))

    recent12 = [v for (_, _, v) in flat[max(0, last_pos - 11):last_pos + 1]]
    out = {
        "status": "ok", "forecast_status": "ok",
        "target_month": target_month, "last_data_month": last_data_month,
        "prev_month": last_data_month, "months_ahead": steps, "steps": steps,
        "generated_at": now,
        "predict": final, "prev": round(lv, 1),
        "delta_pp": delta_pp, "delta_label": "%s 실측 대비" % last_data_month,
        "yoy_pp": yoy_pp, "yoy_prev": (round(yoy_prev, 1) if yoy_prev is not None else None),
        "ci_low": ci_low, "ci_high": ci_high,
        "sigma_base": round(sigma_base or 0.0, 2), "sigma_total": round(total_sigma, 2),
        "path": path, "risk": risk,
        "summary": "", "seasonal_txt": "", "yoy_txt": "", "caution": "", "ai_ok": False,
    }
    ai = _srf_ai(target_month, recent12, tgt_hist, path, steps, out)
    if ai:
        out["ai_ok"] = True
        out["summary"] = ai.get("summary", "") or ""
        out["seasonal_txt"] = ai.get("seasonal", "") or ""
        out["yoy_txt"] = ai.get("yoy", "") or ""
        out["caution"] = ai.get("caution", "") or ""
        if ai.get("risk") in ("상방", "하방", "중립"):
            out["risk"] = ai["risk"]
    return out


if __name__ == "__main__":
    print(json.dumps(compute_sr_forecast(), ensure_ascii=False, indent=2))
