# -*- coding: utf-8 -*-
"""Sea-Intelligence GLP(Global Liner Performance) 보도자료 자동 탐색 + 정시성 수집.

배경: 기존에는 xlsx URL 이 5월호로 하드코딩되어 있어 2026-04 에서 멈춰 있었다.
      GLP 는 M월에 발행하면 M-1월 데이터를 담으므로, 매달 새 URL 을 찾아야 한다.

★ URL 은 규칙 생성이 불가능하다:
    - 앞 숫자(400)는 GLP 전용이 아닌 전체 보도자료 통산 번호
    - 슬러그가 그달 수치·등락에 따라 매번 다름
      (drops-to-62-6-in-june-2026 / for-may-2026-the-highest-of-the-year /
       march-2026-...-joint-highest / 2026-starts-with-... )
  따라서 목록 페이지를 탐색해 최신 GLP 편을 찾는다.

★ 제목 매칭만으로는 부족하다 — GLP 가 아닌 글도 걸린다:
    #401 major-shifts-in-reliability-volatility
    #388 rethinking-the-schedule-reliability-average
  그래서 '제목에 schedule reliability 포함' + '상세 페이지에 press_docs/GLP-*.xlsx 첨부'
  두 조건을 모두 만족하는 것만 GLP 로 인정한다.

★ 파싱 라이브러리: openpyxl (프로젝트에 Node/SheetJS 없음. 기존 SR 파서와 동일 방식).
★ 반환 구조는 기존과 100% 동일하다 — 차트·예측 컴포넌트가 이 키에 의존한다:
    {"status","months":[Jan..Dec],"years":{"2021":[12개],...},"source"}

단독 실행:
    python sea_intelligence.py            # 탐색 + 수집 + 검증 로그
    python sea_intelligence.py --raw      # 엑셀만 tmp/sea-intelligence-raw.xlsx 에 저장 후 종료
    python sea_intelligence.py --dump     # 시트별 첫 20행 출력(구조 확인용)
"""
import io
import os
import re
import sys
import json
import datetime

import requests

SI_ROOT = "https://www.sea-intelligence.com"
SI_PRESS_URL = SI_ROOT + "/press-room"
SI_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
         "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
SI_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# 탐색 실패 시 폴백(기존 하드코딩 URL) — 실패해도 차트가 계속 동작하도록 남겨 둔다.
SI_FALLBACK_XLSX = ("https://www.sea-intelligence.com/images/press_docs/GLP-May2026/"
                    "20260527_-_Sea-Intelligence_GLP_Press_Release_-_May_2026.xlsx")

SI_TITLE_KEY = "schedule reliability"   # 제목 필터(대소문자 무시)
SI_GLP_DOC = "press_docs/glp-"          # 첨부 경로 판별(소문자 비교)
SI_MAX_CANDIDATES = 6                   # 상세 페이지를 열어볼 최대 후보 수

_SI_CACHE = {}                          # 프로세스 내 1회만 탐색(3개 모듈이 공유)


# 직전 데이터이자 폴백 소스. 이 파일 하나가 '기존 상태' 역할을 한다(별도 스냅샷 없음).
SI_DATA_REL = os.path.join("public", "data", "schedule-reliability.json")


def _si_base_dir():
    return os.path.dirname(os.path.abspath(__file__))


def si_data_path():
    """커밋되는 데이터 파일 경로. ★ 읽기만 할 때는 디렉터리를 만들지 않는다
       (배포 번들은 읽기 전용이라 mkdir 자체가 OSError 를 낸다)."""
    return os.path.join(_si_base_dir(), SI_DATA_REL)


def _si_tmp(name):
    """백업 등 부가 용도. 쓰기 불가 환경이면 OS 임시 디렉터리로 폴백."""
    d = os.path.join(_si_base_dir(), "tmp")
    try:
        if not os.path.isdir(d):
            os.makedirs(d)
        return os.path.join(d, name)
    except OSError:
        import tempfile
        return os.path.join(tempfile.gettempdir(), name)


def _si_session():
    s = requests.Session()
    s.headers.update({"User-Agent": SI_UA})
    return s


def _si_log(msg, verbose=True):
    if verbose:
        print("[sea-intel] " + msg)


def _si_slug_title(slug):
    """'400-global-schedule-reliability-drops-to-62-6-in-june-2026'
       → ('400', 'global schedule reliability drops to 62 6 in june 2026')"""
    m = re.match(r"^(\d+)-(.*)$", slug)
    if not m:
        return None, slug.replace("-", " ").lower()
    return m.group(1), m.group(2).replace("-", " ").lower()


def si_list_candidates(sess, verbose=True):
    """목록 페이지에서 제목에 'schedule reliability' 가 든 글을 통산번호 내림차순으로."""
    r = sess.get(SI_PRESS_URL, timeout=30)
    r.raise_for_status()
    html = r.content.decode("utf-8", "replace")
    _si_log("1) 목록 페이지 %s → HTTP %d, %d bytes"
            % (SI_PRESS_URL, r.status_code, len(r.content)), verbose)

    seen, items = set(), []
    for href in re.findall(r'href="(/press-room/[^"#?]+)"', html):
        slug = href.rsplit("/", 1)[-1]
        if slug in seen:
            continue
        seen.add(slug)
        num, title = _si_slug_title(slug)
        items.append({"slug": slug, "url": SI_ROOT + href, "num": int(num) if num else -1,
                      "title": title})
    _si_log("   기사 %d건(중복 제거)" % len(items), verbose)

    hits = [x for x in items if SI_TITLE_KEY in x["title"]]
    hits.sort(key=lambda x: x["num"], reverse=True)
    _si_log("2) 제목에 '%s' 포함: %d건" % (SI_TITLE_KEY, len(hits)), verbose)
    for x in hits[:SI_MAX_CANDIDATES]:
        _si_log("     #%s %s" % (x["num"], x["slug"]), verbose)
    return hits


def si_detail(sess, item, verbose=True):
    """상세 페이지에서 발행일과 xlsx 링크를 뽑는다. GLP 첨부가 없으면 None."""
    r = sess.get(item["url"], timeout=30)
    r.raise_for_status()
    html = r.content.decode("utf-8", "replace")

    hrefs = re.findall(r'href="([^"]+)"', html)
    xlsx = [h for h in hrefs if h.lower().split("?")[0].endswith((".xlsx", ".xls"))]
    # PDF/Word 와 혼동하지 않도록 확장자로만 고른다 + GLP 문서함 경로인지 확인
    glp = [h for h in xlsx if SI_GLP_DOC in h.lower()]
    if not glp:
        return None

    date = None
    m = re.search(r'datetime="([^"]+)"', html)
    if m:
        date = m.group(1)[:10]

    url = glp[0] if glp[0].startswith("http") else SI_ROOT + glp[0]
    others = {e: len([h for h in hrefs if h.lower().split("?")[0].endswith(e)])
              for e in (".pdf", ".docx", ".doc")}
    return {"xlsx": url, "date": date, "other_files": others}


def si_discover(sess=None, verbose=True):
    """최신 GLP 편을 찾아 {title,url,date,xlsx} 반환. 실패 시 None."""
    sess = sess or _si_session()
    try:
        cands = si_list_candidates(sess, verbose)
    except requests.exceptions.RequestException as e:
        _si_log("목록 페이지 실패: %s" % e, verbose)
        return None

    for x in cands[:SI_MAX_CANDIDATES]:
        try:
            d = si_detail(sess, x, verbose)
        except requests.exceptions.RequestException as e:
            _si_log("   #%s 상세 실패(건너뜀): %s" % (x["num"], e), verbose)
            continue
        if not d:
            _si_log("   #%s GLP 엑셀 첨부 없음 → GLP 아님, 건너뜀 (%s)"
                    % (x["num"], x["slug"]), verbose)
            continue
        picked = dict(x)
        picked.update(d)
        _si_log("3) ★ 선택된 보도자료", verbose)
        _si_log("     제목  : %s" % picked["slug"], verbose)
        _si_log("     URL   : %s" % picked["url"], verbose)
        _si_log("     발행일 : %s" % (picked["date"] or "(확인 불가)"), verbose)
        _si_log("4) 엑셀 링크: %s" % picked["xlsx"], verbose)
        _si_log("     (같은 페이지의 pdf/doc 첨부는 제외: %s)"
                % picked["other_files"], verbose)
        return picked

    _si_log("GLP 편을 찾지 못했습니다(후보 %d건 모두 탈락)" % len(cands), verbose)
    return None


def si_latest_xlsx_url(verbose=False):
    """탐색된 최신 GLP xlsx URL(캐시). 실패하면 폴백 URL."""
    if "url" in _SI_CACHE:
        return _SI_CACHE["url"]
    found = si_discover(verbose=verbose)
    url = found["xlsx"] if found else SI_FALLBACK_XLSX
    _SI_CACHE["url"] = url
    _SI_CACHE["meta"] = found
    if not found:
        _si_log("탐색 실패 → 폴백 URL 사용(기존 데이터 유지)", verbose)
    return url


def si_download(url, sess=None, verbose=True):
    sess = sess or _si_session()
    r = sess.get(url, timeout=40, allow_redirects=True)
    r.raise_for_status()
    if r.content[:2] != b"PK":
        raise ValueError("엑셀이 아닌 응답(접근 차단 가능): %d bytes" % len(r.content))
    _si_log("5) 엑셀 다운로드 완료 %d bytes" % len(r.content), verbose)
    return r.content


def si_parse(blob, verbose=True):
    """'Fig 1' 시트 파싱 → {"months":[...], "years":{"2021":[12개], ...}}.
       레이아웃(실측): 1행=[None, Jan..Dec], 2행부터 [연도, 0~1 비율 12개]."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    if "Fig 1" not in wb.sheetnames:
        raise ValueError("'Fig 1' 시트 없음 (시트: %s)" % wb.sheetnames)
    ws = wb["Fig 1"]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    month_cols = [i for i, v in enumerate(header)
                  if isinstance(v, str) and v.strip()[:3] in SI_MONTHS]
    if len(month_cols) < 12:
        month_cols = list(range(1, 13))   # 헤더 파싱 실패 시 관례 위치(B~M)

    years = {}
    for row in ws.iter_rows(min_row=2):
        cells = [c.value for c in row]
        if not cells:
            continue
        try:
            yi = int(cells[0])
        except (TypeError, ValueError):
            continue
        if not (2000 <= yi <= 2100):
            continue
        vals = []
        for ci in month_cols:
            v = cells[ci] if ci < len(cells) else None
            if isinstance(v, (int, float)):
                vals.append(round(v * 100.0 if v <= 1.5 else float(v), 1))
            else:
                vals.append(None)
        years[str(yi)] = vals
    if not years:
        raise ValueError("연도별 데이터 행 없음")
    return {"months": list(SI_MONTHS), "years": years}


def si_latest_period(years):
    """값이 있는 마지막 (연,월) → 'YYYY-MM'."""
    best = None
    for y, vals in years.items():
        for i, v in enumerate(vals):
            if v is not None:
                p = "%s-%02d" % (y, i + 1)
                if best is None or p > best:
                    best = p
    return best


def si_count_values(years):
    return sum(1 for vals in years.values() for v in vals if v is not None)


def si_validate(parsed, verbose=True):
    """0~100% 범위 확인. 벗어나면 파싱 오류로 보고 예외."""
    bad = []
    for y, vals in parsed["years"].items():
        for i, v in enumerate(vals):
            if v is None:
                continue
            if not (0.0 <= v <= 100.0):
                bad.append("%s-%02d=%s" % (y, i + 1, v))
    if bad:
        raise ValueError("정시성 값이 0~100%% 범위를 벗어남: %s" % ", ".join(bad[:6]))
    _si_log("   범위 검증 통과(0~100%%) · 연도 %d개 / 값 %d개"
            % (len(parsed["years"]), si_count_values(parsed["years"])), verbose)


def si_load_existing(verbose=False):
    """커밋된 public/data/schedule-reliability.json 을 '직전 상태'로 읽는다.
       ★ 절대 예외를 올리지 않는다 — 없거나 못 읽어도 수집을 막으면 안 된다.
       ★ 디렉터리를 만들지 않는다(읽기 전용 배포 번들 대응)."""
    try:
        p = si_data_path()
        if not os.path.isfile(p):
            return None
        data = json.load(io.open(p, encoding="utf-8"))
        if isinstance(data.get("years"), dict) and data["years"]:
            return data
        return None
    except Exception as e:  # noqa: BLE001
        _si_log("   기존 데이터 읽기 실패(무시하고 진행): %s" % e, verbose)
        return None


def si_save_data(payload, verbose=True):
    """같은 경로에 덮어쓴다. 교체 전 기존 파일을 tmp/ 에 백업.
       ★ 쓰기 실패(배포 번들=읽기 전용)해도 수집 결과에는 영향을 주지 않는다."""
    p = si_data_path()
    try:
        if os.path.isfile(p):
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            bak = _si_tmp("schedule-reliability-%s.json" % stamp)
            try:
                with io.open(bak, "w", encoding="utf-8") as f:
                    f.write(io.open(p, encoding="utf-8").read())
                _si_log("   기존 데이터 백업: %s" % bak, verbose)
            except Exception as e:  # noqa: BLE001
                _si_log("   백업 실패(계속 진행): %s" % e, verbose)
        d = os.path.dirname(p)
        if not os.path.isdir(d):
            os.makedirs(d)
        with io.open(p, "w", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False, indent=1))
        _si_log("   저장: %s" % p, verbose)
        return True
    except Exception as e:  # noqa: BLE001
        # Vercel 등 읽기 전용 환경에서는 정상적인 상황이다(런타임 수집 결과는 그대로 반환).
        _si_log("   저장 생략(쓰기 불가 환경): %s" % e, verbose)
        return False


def update_schedule_reliability_auto(verbose=True):
    """최신 GLP 자동 탐색 → 수집 → 검증 → 기존 구조 그대로 반환.

    흐름: 기존 public/data JSON 읽기 → 새로 수집 → 범위·행 수 검증 →
          통과하면 같은 경로에 덮어쓰기 → 실패하면 기존 파일 유지.
    ★ 파일 입출력은 전부 best-effort 다. 쓰기 불가 환경(Vercel 배포 번들)에서도
      수집 결과는 그대로 반환된다 — 예전엔 여기서 OSError 가 나 폴백으로 빠졌다.
    """
    prev = si_load_existing(verbose)          # 예외를 올리지 않음

    # ── 수집·파싱만 오류로 취급한다(부가 기능 실패로 폴백되지 않게) ──
    try:
        sess = _si_session()
        found = si_discover(sess, verbose)
        url = found["xlsx"] if found else SI_FALLBACK_XLSX
        if not found:
            _si_log("탐색 실패 → 폴백 URL 로 시도: %s" % url, verbose)
        blob = si_download(url, sess, verbose)
        parsed = si_parse(blob, verbose)
        si_validate(parsed, verbose)          # 0~100% 이탈 시 여기서 중단
    except Exception as e:  # noqa: BLE001  (RequestException 포함)
        if prev:
            _si_log("수집 실패(%s) → 커밋된 기존 데이터 유지: 최신 %s"
                    % (e, prev.get("latest_period")), verbose)
            out = dict(prev)
            out["status"] = "ok"
            out["warning"] = "수집 실패로 기존 데이터 사용: %s" % e
            return out
        return {"status": "error", "reason": "수집/파싱 실패: %s" % e}

    latest = si_latest_period(parsed["years"])
    n_new = si_count_values(parsed["years"])

    # 안전장치: 값 개수가 기존 파일보다 적으면 교체하지 않는다.
    # ★ 기존 파일이 없는 최초 실행에서는 가드를 건너뛰고 그냥 저장한다.
    if prev:
        n_old = si_count_values(prev["years"])
        if n_new < n_old:
            _si_log("⚠ 새 데이터가 기존보다 적음(%d < %d) → 교체하지 않고 기존 유지"
                    % (n_new, n_old), verbose)
            out = dict(prev)
            out["status"] = "ok"
            out["warning"] = ("새 데이터 값 %d개가 기존 %d개보다 적어 교체하지 않음"
                              % (n_new, n_old))
            return out
        _si_log("   값 개수 %d → %d (기존 대비 %+d)" % (n_old, n_new, n_new - n_old), verbose)
    else:
        _si_log("   기존 데이터 없음(최초 실행) → 가드 생략하고 저장", verbose)

    payload = {
        "status": "ok",
        "months": parsed["months"],
        "years": parsed["years"],
        "source": "Sea-Intelligence",
        "press_title": (found or {}).get("slug"),
        "press_url": (found or {}).get("url"),
        "press_date": (found or {}).get("date"),
        "xlsx_url": url,
        "latest_period": latest,
        "updatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    si_save_data(payload, verbose)            # 실패해도 payload 는 그대로 반환
    _si_log("최신 데이터: %s" % latest, verbose)
    return payload


def _si_dump(blob):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    print("[sea-intel] 시트 목록: %s" % wb.sheetnames)
    for name in wb.sheetnames:
        print("[sea-intel] ── 시트 '%s' 첫 20행 ──" % name)
        for i, row in enumerate(wb[name].iter_rows(min_row=1, max_row=20, values_only=True)):
            cells = ["" if c is None else (round(c, 4) if isinstance(c, float) else c) for c in row]
            if any(str(c) != "" for c in cells):
                print("    r%-2d %s" % (i + 1, cells[:15]))


if __name__ == "__main__":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

    if "--raw" in sys.argv or "--dump" in sys.argv:
        sess = _si_session()
        found = si_discover(sess, True)
        url = found["xlsx"] if found else SI_FALLBACK_XLSX
        blob = si_download(url, sess, True)
        path = _si_tmp("sea-intelligence-raw.xlsx")
        open(path, "wb").write(blob)
        print("[sea-intel] --raw 저장: %s (%d bytes)" % (path, len(blob)))
        if "--dump" in sys.argv:
            _si_dump(blob)
        sys.exit(0)

    res = update_schedule_reliability_auto(True)
    if res.get("status") != "ok":
        print("[sea-intel] 실패: %s" % res.get("reason"))
        sys.exit(1)
    print(json.dumps({k: v for k, v in res.items() if k != "years"},
                     ensure_ascii=False, indent=2))
    for y in sorted(res["years"]):
        print("  %s %s" % (y, res["years"][y]))
