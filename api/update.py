# -*- coding: utf-8 -*-
"""
SIMMONS 대시보드 — Vercel 서버리스 함수 (/api/update).

요청(GET/POST)이 오면 아래 무료·무키 소스에서 데이터를 모아 JSON으로 반환한다:
  - materials   : World Bank 핑크시트(월간 원자재 가격, xlsx)
  - market      : World Bank Open Data (수요 잠재력 지표)
  - news        : Google News RSS + 제목 한국어 번역
  - competitors : SEC EDGAR (SNBR·TPX 재무/매출추이)
  - fx          : Frankfurter(ECB) USD/KRW 환율

응답 형식은 로컬용 dashboard_server.py 의 /api/update 와 동일하다:
  { "updated_at": "...", "sections": { "materials": {...}, ... } }

※ 이 파일은 dashboard_server.py 의 수집 로직을 그대로 옮긴 "배포용" 사본이다.
  로직을 고칠 때는 두 파일을 함께 맞춰야 한다(로컬 테스트는 dashboard_server.py 사용).
"""
import re
import os
import sys
import io
import json
import time
import datetime
import urllib.parse
import concurrent.futures
import xml.etree.ElementTree as ET
from http.server import BaseHTTPRequestHandler

import requests

# 순수 추가: 다음 달 주원료 예측 모듈(프로젝트 루트). 로드 실패해도 나머지는 정상.
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from icis_forecast import compute_icis_forecast
except Exception:  # noqa: BLE001
    compute_icis_forecast = None
try:
    from sr_forecast import compute_sr_forecast
except Exception:  # noqa: BLE001
    compute_sr_forecast = None
try:
    from oil_forecast import compute_oil_forecast
except Exception:  # noqa: BLE001
    compute_oil_forecast = None
try:  # 순수 추가: KOIMA 월간 부문별 지수
    from koima_index import update_koima_index
except Exception:  # noqa: BLE001
    update_koima_index = None
try:  # 순수 추가: 해상 정시성 최신 GLP 자동 탐색(실패 시 기존 고정 URL 로 폴백)
    from sea_intelligence import update_schedule_reliability_auto
except Exception:  # noqa: BLE001
    update_schedule_reliability_auto = None


# ── 시몬스 코리아 소식 — Google News RSS (API 키 불필요) ─────────────────
# 정밀 쿼리(따옴표+OR)로 관련성 높은 기사만. 결과 부족 시 '시몬스' 광의 검색으로 폴백.
SIMMONS_NARROW_QUERY = '"시몬스 침대" OR "시몬스코리아" OR "시몬스 그로서리"'
SIMMONS_FALLBACK_QUERY = "시몬스"
# 제목에 이 브랜드가 함께 있으면 '타 브랜드가 주어'일 수 있어 애매 → 통과시키되 로그로 표시
SIMMONS_OTHER_BRANDS = ["세라젬", "쿤달", "코웨이", "바디프랜드", "에이스침대", "에이스 침대",
                        "씰리", "한샘", "이케아", "템퍼", "슬립앤슬립", "라클라우드",
                        "알레르망", "이브자리", "지누스", "웰스"]
_GNEWS_UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}
# 기사 페이지에서 대표 이미지 후보 (og:image → twitter:image → link[image_src])
_IMG_METAS = [
    re.compile(r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]+content=["\']([^"\']+)', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']', re.I),
    re.compile(r'<meta[^>]+name=["\']twitter:image(?::src)?["\'][^>]+content=["\']([^"\']+)', re.I),
    re.compile(r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)', re.I),
]


def _decode_gnews_url(link, timeout=8):
    """Google News RSS 리다이렉트 링크 → 실제 기사 URL(batchexecute). 실패 시 원 링크."""
    try:
        from bs4 import BeautifulSoup
        m = re.search(r'/articles/([^?/]+)', link)
        if not m:
            return link
        gid = m.group(1)
        r = requests.get("https://news.google.com/rss/articles/" + gid, headers=_GNEWS_UA, timeout=timeout)
        r.raise_for_status()
        div = BeautifulSoup(r.text, "html.parser").select_one("c-wiz > div")
        if div is None:
            return link
        sig, ts = div.get("data-n-a-sg"), div.get("data-n-a-ts")
        if not sig or not ts:
            return link
        inner = json.dumps(["garturlreq",
                            [["X", "X", ["X", "X"], None, None, 1, 1, "US:en", None, 1,
                              None, None, None, None, None, 0, 1],
                             "X", "X", 1, [1, 1, 1], 1, 1, None, 0, 0, None, 0],
                            gid, int(ts), sig])
        freq = json.dumps([[["Fbv4je", inner, None, "generic"]]])
        resp = requests.post("https://news.google.com/_/DotsSplashUi/data/batchexecute",
                             headers={"Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
                                      **_GNEWS_UA},
                             data={"f.req": freq}, timeout=timeout)
        # 응답은 이중 이스케이프된 JSON 문자열 — \\ 축소 후 \\uXXXX(=,& 등) 디코드
        m2 = re.search(r'garturlres\\",\\"(.*?)\\"', resp.text)
        if m2:
            u = m2.group(1).replace("\\\\", "\\")
            u = re.sub(r"\\u([0-9a-fA-F]{4})", lambda mm: chr(int(mm.group(1), 16)), u)
            u = u.replace("\\/", "/").replace("\\", "")  # 남은 escape 백슬래시 제거
            if u.startswith("http"):
                return u
    except Exception:  # noqa: BLE001
        pass
    return link


def _page_image(page_url, timeout=8):
    """기사 페이지 대표 이미지: og:image → twitter:image → link[image_src] → 본문 첫 큰 이미지."""
    if not page_url or not page_url.startswith("http"):
        return None
    try:
        r = requests.get(page_url, headers=_GNEWS_UA, timeout=timeout)
        r.raise_for_status()
        html = r.text
    except Exception:  # noqa: BLE001
        return None
    for rx in _IMG_METAS:
        m = rx.search(html)
        if m:
            img = urllib.parse.urljoin(page_url, m.group(1).strip())
            if img.startswith("http") and "googleusercontent" not in img:
                return img
    for m in re.finditer(r'<img[^>]+(?:data-src|src)=["\']([^"\']+\.(?:jpe?g|png|webp)[^"\']*)', html, re.I):
        img = urllib.parse.urljoin(page_url, m.group(1).strip())
        if img.startswith("http") and "googleusercontent" not in img:
            return img
    return None


def _simmons_thumb(link):
    """Google News 링크 → 실제 기사 URL 디코드 → 대표 이미지. 실패 시 None."""
    try:
        return _page_image(_decode_gnews_url(link))
    except Exception:  # noqa: BLE001
        return None


def _norm_title(title, source):
    """제목에서 ' - 언론사' 접미사 제거 → (표시용 제목, 중복판정 키: 공백·기호 제거 앞 18자)."""
    t = (title or "").strip()
    if source and t.endswith(" - " + source):
        t = t[:-(len(source) + 3)]
    else:
        t = re.sub(r"\s*-\s*[^-]+$", "", t).strip()
    key = re.sub(r"[\s\W]+", "", t)[:18]
    return t, key


def _gnews_candidates(query):
    """Google News RSS 검색 → 후보 리스트[{title,link,source,date}] (제목 중복 제거)."""
    url = ("https://news.google.com/rss/search?q=" + urllib.parse.quote(query, safe="")
           + "&hl=ko&gl=KR&ceid=KR:ko")  # 따옴표·공백·한글 모두 인코딩(encodeURIComponent 상당)
    out, seen = [], set()
    try:
        root = ET.fromstring(requests.get(url, timeout=20, headers=_GNEWS_UA).content)
    except Exception:  # noqa: BLE001
        return out
    for item in root.iter("item"):
        raw = (item.findtext("title") or "").strip()
        if not raw:
            continue
        src_el = item.find("source")
        source = (src_el.text.strip() if (src_el is not None and src_el.text) else "")
        disp, key = _norm_title(raw, source)
        if not key or key in seen:
            continue
        seen.add(key)
        desc = re.sub(r"<[^>]+>", " ", item.findtext("description") or "")  # RSS 스니펫(태그 제거)
        desc = re.sub(r"&[a-zA-Z#0-9]+;", " ", desc)
        desc = re.sub(r"\s+", " ", desc).strip()
        out.append({"title": disp, "link": (item.findtext("link") or "").strip(),
                    "source": source, "date": _fmt_pubdate(item.findtext("pubDate") or ""),
                    "desc": desc})
    return out


def _simmons_relevant(cands):
    """제목 기준 관련성 필터. 반환 (kept, excluded_titles, flagged_titles).
       - 제목에 '시몬스'/'simmons' 없으면 제외(요약에만 스친 기사 배제).
       - 타 브랜드가 함께 언급되면 애매 → 통과시키되 flagged로 기록(사람 확인용)."""
    kept, excluded, flagged = [], [], []
    for it in cands:
        t = it["title"]
        if ("시몬스" not in t) and ("simmons" not in t.lower()):
            excluded.append(t)
            continue
        others = [b for b in SIMMONS_OTHER_BRANDS if b in t]
        if others:
            flagged.append("%s  ← 타 브랜드 언급(%s)" % (t, ",".join(others)))
        kept.append(it)
    return kept, excluded, flagged


# 나열형(종합) 기사 판별용
SIMMONS_LISTICLE_PHRASES = ["새로나왔어요", "새로 나왔어요", "신제품 모음", "신상품 모음",
                            "이번 주 신제품", "이주의 신제품", "한눈에"]
SIMMONS_ROUNDUP_SECTIONS = ["유통", "신상", "신제품", "새상품", "리테일", "이주의", "쇼핑",
                            "픽", "pick", "브리핑", "하이라이트"]
# 시몬스와 무관한(다른 업종) 기업/브랜드 — 제목에 2개↑면 나열형 종합 기사로 판단
SIMMONS_OTHER_COMPANIES = [
    "서울우유", "남양유업", "매일유업", "빙그레", "오뚜기", "농심", "삼양", "팔도", "동원", "풀무원",
    "대상", "하이트진로", "롯데칠성", "롯데웰푸드", "코카콜라", "펩시", "맥도날드", "한국맥도날드",
    "버거킹", "롯데리아", "kfc", "맘스터치", "스타벅스", "투썸", "이디야", "메가커피", "컴포즈", "빽다방",
    "파리바게뜨", "뚜레쥬르", "배스킨라빈스", "던킨", "공차", "cu", "gs25", "세븐일레븐", "이마트24",
    "홈플러스", "올리브영", "다이소", "무신사", "컬리"]


def _simmons_topic_filter(cands, body_check=True):
    """시몬스가 '주제'인지 판정. 반환 (kept, log[(title, ok, reason)]).
       1) 나열형(종합) 기사 제외: 外/외, 신제품 모음 표현, 종합 코너 대괄호, 타 기업 2개↑ 나열
       2) 본문(RSS description) 기반: 제목 앞 15자에 시몬스 있으면 통과,
          뒤에만 있으면 요약의 '시몬스' 언급이 2회 미만이면 제외(요약 없으면 판정보류·통과)."""
    kept, log = [], []
    for it in cands:
        t = it["title"]
        tl = t.lower()
        reason = None
        mbr = re.match(r"^\s*\[([^\]]*)\]", t)  # 선두 대괄호 섹션명
        if ("外" in t) or re.search(r"외\s*\(", t) or re.search(r"외\s*\d+\s*건", t):
            reason = "나열형(外/외 나열)"
        elif any(p in t for p in SIMMONS_LISTICLE_PHRASES):
            reason = "나열형(신제품 모음 표현)"
        elif mbr and any(k in mbr.group(1).lower() for k in SIMMONS_ROUNDUP_SECTIONS):
            reason = "종합 코너 대괄호([%s])" % mbr.group(1).strip()
        else:
            others = sorted({c for c in SIMMONS_OTHER_COMPANIES if c in tl})
            if len(others) >= 2:
                reason = "타 기업 %d개 나열(%s)" % (len(others), ",".join(others))
        if reason:
            log.append((t, False, reason))
            continue

        if not body_check:  # 완화 모드: 나열형만 거르고 통과
            kept.append(it)
            log.append((t, True, "통과(본문판정 완화)"))
            continue

        pos = t.find("시몬스")
        if pos < 0:
            pos = tl.find("simmons")
        if 0 <= pos < 15:  # 제목 앞부분이 시몬스 → 주제 가능성 높음
            kept.append(it)
            log.append((t, True, "제목 앞부분 주제"))
            continue
        desc = it.get("desc") or ""
        if not desc:  # 요약 없음 → 판정 불가, 안전하게 통과
            kept.append(it)
            log.append((t, True, "판정보류(요약 없음)"))
            continue
        cnt = desc.count("시몬스") + desc.lower().count("simmons")
        if cnt < 2:
            log.append((t, False, "본문(요약) 시몬스 언급 %d회(<2)" % cnt))
            continue
        kept.append(it)
        log.append((t, True, "본문 시몬스 언급 %d회" % cnt))
    return kept, log


def _bigrams(s):
    if len(s) >= 2:
        return set(s[i:i + 2] for i in range(len(s) - 1))
    return {s} if s else set()


def _title_jaccard(a, b):
    """2-gram Jaccard 유사도(0~1)."""
    A, B = _bigrams(a), _bigrams(b)
    if not A and not B:
        return 1.0
    if not A or not B:
        return 0.0
    return len(A & B) / len(A | B)


def _set_jaccard(a, b):
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


# 표기 정규화 사전(영문↔한글 음차) · 동의어 통일 · 흔한 수식어 stopword
WC_NOTATION = [
    ("cool summer", "쿨썸머"), ("쿨 썸머", "쿨썸머"),
    ("life is comfort", "라이프이즈컴포트"), ("라이프 이즈 컴포트", "라이프이즈컴포트"),
    ("beautyrest black", "뷰티레스트블랙"), ("뷰티레스트 블랙", "뷰티레스트블랙"),
    ("refresh & rewards", "리프레시리워즈"), ("bigger & more", "비거앤모어"), ("비거 앤 모어", "비거앤모어"),
]
WC_SYNONYMS = [
    (["기획전", "프로모션", "행사", "이벤트", "페어", "세일"], "PROMO"),
    (["실시", "전개", "진행", "개최", "운영", "오픈", "선사", "제공"], "RUN"),
    (["출시", "선봬", "선보여", "선보인다", "선보였다", "선봤다", "공개", "론칭", "런칭"], "LAUNCH"),
]
WC_STOP = set(("여름 봄 가을 겨울 맞이 맞아 맞아서 맞은 위한 위해 관련 기념 신규 새 전국 국내 해외 "
               "오늘 이번 최대 첫 더 및 등 통해 통한 대한 앞두고 함께 침대 소식 시즌 잡는 겨냥 더위 "
               "방문객 시장").split())


def _wc_norm_text(t):
    """소문자화 → 표기 사전(영문 음차) → 동의어 태그 치환."""
    t = (t or "").lower()
    for a, b in WC_NOTATION:
        t = t.replace(a, b)
    for words, tag in WC_SYNONYMS:
        for w in words:
            t = t.replace(w, tag)
    return t


def _wc_strip_simmons(t):
    """모든 기사 공통인 시몬스 표기 제거(그로서리·뷰티레스트 등 실질 키워드는 보존)."""
    t = re.sub(r"시몬스\s*침대", " ", t)
    t = re.sub(r"시몬스\s*테라스", " ", t)
    t = t.replace("시몬스", " ")
    return re.sub(r"simmons", " ", t, flags=re.I)


def _wc_strip_lead_subtitle(t):
    """제목 맨 앞의 따옴표/대괄호로 묶인 부제·코너명 제거."""
    return re.sub(r"^\s*[\"'“”‘’\[(][^\"'“”‘’\)\]]{0,40}[\"'“”‘’\)\]]\s*", "", t or "")


def _wc_keywords(title):
    """제목 → 핵심 키워드 집합(시몬스·부제·조사/수식어 제거, 표기/동의어 정규화)."""
    t = _wc_norm_text(_wc_strip_simmons(_wc_strip_lead_subtitle(title)))
    out = set()
    for w in re.findall(r"[A-Za-z]+|[0-9]+|[가-힣]+", t):
        if w in ("PROMO", "RUN", "LAUNCH"):
            out.add(w)
        elif not w.isdigit() and len(w) >= 2 and w not in WC_STOP:
            out.add(w)
    return out


def _wc_strkey(title):
    """문자열 유사도용: 시몬스 제거 + 표기/동의어 정규화 + 공백·구두점 제거."""
    t = _wc_norm_text(_wc_strip_simmons(_wc_strip_lead_subtitle(title)))
    return re.sub(r"[^0-9a-z가-힣]", "", t)


def _img_norm(u):
    """이미지 URL 정규화: 쿼리·프래그먼트 제거."""
    if not u:
        return None
    return u.split("?")[0].split("#")[0].rstrip("/").lower()


def _rep_key(it):
    """그룹 대표 선정 우선순위: 이미지 있음 → 발행일 빠름 → 제목 짧음."""
    return (0 if it.get("image") else 1, it.get("date") or "", len(it["title"]))


def _dup_match(a, b, kw_thr, str_thr):
    """두 기사 중복 판정. 반환 (matched, rule, score). 우선순위: 이미지(동일 URL) > 키워드 > 문자열.
       ※ 파일명만 비교는 Google News 이미지 프록시 특성상 서로 다른 기사가 충돌(오결합)해 제외."""
    if a.get("_imn") and b.get("_imn") and a["_imn"] == b["_imn"]:
        return True, "이미지(URL)", 1.0
    kj = _set_jaccard(a["_kw"], b["_kw"])
    if kj >= kw_thr:
        return True, "키워드", round(kj, 2)
    sj = _title_jaccard(a["_sk"], b["_sk"])
    if sj >= str_thr:
        return True, "문자열", round(sj, 2)
    return False, None, round(max(kj, sj), 2)


def _cluster_articles(items, kw_thr=0.5, str_thr=0.7, same_fn=None):
    """기사 리스트를 이미지/키워드/문자열 신호로 그리디 군집(공용 — 시몬스·브랜드 섹션 공통).
       same_fn(a, b)=False면 같은 그룹 후보에서 제외(예: 다른 브랜드끼리는 묶지 않음).
       각 item에 _kw/_sk/_imn 부여. 반환 groups=[{items, merges}]."""
    for it in items:
        it["_kw"] = _wc_keywords(it["title"])
        it["_sk"] = _wc_strkey(it["title"])
        it["_imn"] = _img_norm(it.get("image"))
    groups = []
    for it in items:
        placed = False
        for g in groups:
            head = g["items"][0]
            if same_fn and not same_fn(it, head):
                continue
            matched, rule, score = _dup_match(it, head, kw_thr, str_thr)
            if matched:
                g["items"].append(it)
                g["merges"].append((it["title"], rule, score))
                placed = True
                break
        if not placed:
            groups.append({"items": [it], "merges": []})
    return groups


def _simmons_dedup(cands, kw_thr=0.5, str_thr=0.7):
    """2단계 중복 제거.
       1) 텍스트(키워드/문자열)로 저비용 군집 — 풀을 넓게(다양한 최신 주제 확보).
       2) 상위 그룹만 이미지 확보 → 대표 선정(이미지>날짜>길이) + 동일 이미지 URL 그룹 2차 병합.
       → 매체 다양성(매체당 1건) → 6건. 반환 (chosen, groups[{items,merges,rep}])."""
    pool = cands[:40]
    groups = _cluster_articles(pool, kw_thr, str_thr)  # 텍스트 신호(이미지 아직 없음)
    groups.sort(key=lambda g: max((x["date"] or "") for x in g["items"]), reverse=True)  # 최신 그룹 먼저
    head = groups[:12]  # 표시 6 + 이미지병합/매체다양성 여유
    # 상위 그룹의 (이른 날짜·짧은 제목) 상위 2건만 이미지 확보(비용 제한)
    to_img, seen = [], set()
    for g in head:
        for m in sorted(g["items"], key=lambda x: ((x["date"] or ""), len(x["title"])))[:2]:
            if m["link"] and m["link"] not in seen:
                seen.add(m["link"])
                to_img.append(m)
    if to_img:
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(16, len(to_img))) as ex:
                for m, img in zip(to_img, ex.map(lambda it: _simmons_thumb(it["link"]), to_img)):
                    m["image"] = img
        except Exception:  # noqa: BLE001
            pass
    for m in to_img:
        m["_imn"] = _img_norm(m.get("image"))
    for g in head:
        g["rep"] = sorted(g["items"], key=_rep_key)[0]  # 대표: 이미지>날짜>길이
    # 동일 이미지 URL 그룹 2차 병합(문자열론 못 잡는 같은 보도자료)
    merged = []
    for g in head:
        rep, done = g["rep"], False
        imn = rep.get("_imn")
        if imn:
            for mg in merged:
                if mg["rep"].get("_imn") == imn:
                    mg["items"] += g["items"]
                    mg["merges"].append((rep["title"], "이미지(URL)", 1.0))
                    done = True
                    break
        if not done:
            merged.append(g)
    chosen, used_src, seen_img = [], set(), set()
    for g in merged:  # 매체 다양성(매체당 1건) + 이미지 중복 회피
        rep = g["rep"]
        src = rep.get("source") or ""
        imn = rep.get("_imn")
        if src and src in used_src:
            continue
        if imn and imn in seen_img:
            continue
        if src:
            used_src.add(src)
        if imn:
            seen_img.add(imn)
        chosen.append(rep)
        if len(chosen) >= 6:
            break
    if len(chosen) < 6:  # 부족 시 매체 제약 완화(이미지 중복만 회피)
        picked = {id(x) for x in chosen}
        for g in merged:
            rep = g["rep"]
            if id(rep) in picked:
                continue
            imn = rep.get("_imn")
            if imn and imn in seen_img:
                continue
            if imn:
                seen_img.add(imn)
            chosen.append(rep)
            if len(chosen) >= 6:
                break
    return chosen, merged


def _simmons_plain_top(cands, n):
    """중복제거 생략 폴백 — 최신순 상위 n건(+이미지, 이미지 중복만 회피)."""
    top = cands[:max(n + 3, 9)]
    imgs = [None] * len(top)
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(16, len(top) or 1)) as ex:
            imgs = list(ex.map(lambda it: _simmons_thumb(it["link"]), top))
    except Exception:  # noqa: BLE001
        pass
    out, seen_img = [], set()
    for it, img in zip(top, imgs):
        if img and img in seen_img:
            continue
        if img:
            seen_img.add(img)
        it["image"] = img
        out.append(it)
        if len(out) >= n:
            break
    return out


def update_simmons_news():
    """정밀 쿼리 수집 → 관련성 필터 → 제목 유사도 중복 제거(매체 다양성) → 최신순 6건.
       필터/중복제거 후 부족하면 단계적 완화. 검증 로그는 서버 콘솔에 출력."""
    cands = _gnews_candidates(SIMMONS_NARROW_QUERY)
    pre = len(cands)
    kept, excluded, flagged = _simmons_relevant(cands)

    used_fallback = False
    if len(kept) < 3:  # 관련성 필터가 빡빡해 부족하면 광의 검색으로 보충
        used_fallback = True
        seen_keys = {_norm_title(x["title"], x["source"])[1] for x in kept}
        fb_kept, fb_ex, fb_fl = _simmons_relevant(_gnews_candidates(SIMMONS_FALLBACK_QUERY))
        for it in fb_kept:
            k = _norm_title(it["title"], it["source"])[1]
            if k not in seen_keys:
                seen_keys.add(k)
                kept.append(it)
        excluded += fb_ex
        flagged += fb_fl
        print("[simmons_news] WARN 관련성 필터 후 3건 미만 → '시몬스' 광의 검색 폴백")

    kept.sort(key=lambda x: x["date"], reverse=True)  # 최신순

    print("[simmons_news] 수집 원본 %d건 → 관련성 필터 후 %d건%s"
          % (pre, len(kept), " (폴백 포함)" if used_fallback else ""))
    if excluded:
        print("[simmons_news] 관련성 제외 %d건:\n  - %s" % (len(excluded), "\n  - ".join(excluded)))
    if flagged:
        print("[simmons_news] * 애매(사람 확인 필요) %d건:\n  - %s" % (len(flagged), "\n  - ".join(flagged)))
    if not kept:
        print("[simmons_news] 최종 0건 — 데이터 없음")
        return {"status": "최근 관련 기사가 없습니다", "items": []}

    # 주제 적합성 필터(나열형 제외 + 본문 기반). 3건 미만이면 본문 판정만 완화(나열형은 유지).
    topic_kept, topic_log = _simmons_topic_filter(kept, body_check=True)
    if len(topic_kept) < 3:
        print("[simmons_news] WARN 주제 필터 후 3건 미만 → 본문 판정 완화(나열형 규칙만 유지)")
        topic_kept, topic_log = _simmons_topic_filter(kept, body_check=False)
    print("[simmons_news] 주제 판정 (통과 %d / 검사 %d):" % (len(topic_kept), len(topic_log)))
    for title, ok, reason in topic_log:
        print("   [%s] %s — %s" % ("통과" if ok else "제외", title, reason))
    if not topic_kept:
        print("[simmons_news] 주제 적합 0건 — 표시할 기사 없음")
        return {"status": "최근 관련 기사가 없습니다", "items": []}

    # 중복 제거: 이미지 URL > 키워드집합 Jaccard(≥0.5) > 문자열 유사도(≥0.7).
    # 부족 시 임계값 완화(키워드 0.65/문자열 0.8), 그래도 부족하면 중복제거 생략.
    chosen, groups = _simmons_dedup(topic_kept)
    if len(chosen) < 3:
        print("[simmons_news] WARN 중복제거 후 3건 미만 → 임계값 완화(키워드 0.65/문자열 0.8) 재계산")
        chosen, groups = _simmons_dedup(topic_kept, kw_thr=0.65, str_thr=0.8)
    if len(chosen) < 3:
        print("[simmons_news] WARN 완화 후에도 부족 → 중복제거 생략, 최신순 상위 6건 표시")
        chosen, groups = _simmons_plain_top(topic_kept, 6), None

    # ── 중복 그룹 로그(2건 이상 + 묶인 규칙/유사도 + 대표) ──
    if groups is not None:
        multi = [g for g in groups if len(g["items"]) > 1]
        print("[simmons_news] 중복 그룹 %d개(2건+):" % len(multi))
        for gi, g in enumerate(multi, 1):
            items_g = g["items"]
            rep = sorted(items_g, key=_rep_key)[0]
            warn = " ⚠4건 초과(과결합 여부 확인)" if len(items_g) > 4 else ""
            print("  [그룹 %d] %d건%s — 대표: %s" % (gi, len(items_g), warn, rep["title"]))
            how = {t: (r, s) for (t, r, s) in g["merges"]}
            for m in items_g:
                rs = how.get(m["title"])
                via = ("  ← %s %.2f" % (rs[0], rs[1])) if rs else "  (기준)"
                print("     · %s%s%s" % (m["title"], via, "   ★대표" if m is rep else ""))

    items = [{"title": c["title"], "link": c["link"], "source": c["source"],
              "date": c["date"], "image": c.get("image")} for c in chosen[:6]]
    print("[simmons_news] 최종 표시 %d건:\n  - %s"
          % (len(items), "\n  - ".join(x["title"] for x in items)))
    return {"status": "ok", "items": items}


def _fmt_pubdate(s):
    """RFC822 pubDate → 'YYYY-MM-DD'."""
    s = (s or "").strip()
    for fmt in ("%a, %d %b %Y %H:%M:%S %Z", "%a, %d %b %Y %H:%M:%S %z"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:  # noqa: BLE001
            pass
    return s[:16]


# ── 국내외 브랜드 신제품 — Google News RSS (무료·무키) ───────────────────
# 국내: 상단(대표 출시 상품)과 하단(기사 목록)이 서로 다른 브랜드를 본다.
# 수집은 두 목록의 합집합으로 '한 번만' 하고, 표시 단계에서 각각 필터링한다.
DOMESTIC_FEATURED_BRANDS = ["한샘", "에이스침대", "이케아"]                    # 상단 3칸
DOMESTIC_ITEM_BRANDS = ["씰리침대", "템퍼", "슬럼버랜드", "킹스다운", "지누스"]   # 하단 목록
DOMESTIC_BRANDS = DOMESTIC_FEATURED_BRANDS + DOMESTIC_ITEM_BRANDS            # 수집 대상(합집합)

# 국내 브랜드 검색 쿼리 — 이름이 일반 단어·타 브랜드와 겹치면 '침대/매트리스'를 함께 요구한다.
# (실측 충돌: "템퍼" → 리더뮨 '템퍼픽션'(구강용해필름) 기사,
#             "슬럼버랜드" → 애니메이션 영화 Slumberland 기사)
# ★ 값이 있으면 그 문자열을 그대로 쓴다(신제품 절을 붙이지 않는다).
#   슬럼버랜드·킹스다운은 국내 기사 수 자체가 적어 신제품 절을 붙이면 0건이 되므로 뺐다.
#   대신 하류의 _news_grade 가 실적·M&A 기사를 걸러 준다.
DOMESTIC_BRAND_QUERY = {
    "씰리침대": '("씰리침대" OR "씰리 매트리스" OR "Sealy") (신제품 OR 출시 OR 론칭)',
    "템퍼": '("템퍼 매트리스" OR "템퍼코리아" OR "템퍼 침대" OR "템퍼 모션베드" OR "Tempur")'
            ' (신제품 OR 출시 OR 론칭)',
    "슬럼버랜드": '("슬럼버랜드 매트리스" OR "슬럼버랜드 침대" OR "슬럼버랜드 호텔")',
    "킹스다운": '("킹스다운 매트리스" OR "킹스다운 침대")',
    "지누스": '("지누스" OR "지누스 매트리스" OR "Zinus") (신제품 OR 출시 OR 론칭)',
}
GLOBAL_BRANDS = ["Sleep Number", "Tempur-Pedic", "Purple", "Serta"]  # 국외(영어 검색)

# 이름이 일반 단어·타 고유명사와 충돌하는 브랜드는 정확 구문으로 좁힌다.
# (기본 쿼리 '"브랜드" (new OR launch OR release OR mattress)' 는 OR 때문에
#  'Purple' + 'new' 만으로도 걸려 록밴드 Deep Purple 기사가 섞였다)
# ★ Casper/Nectar/Saatva/Sealy 는 현재 GLOBAL_BRANDS 에 없지만, 나중에 추가해도
#   바로 안전하도록 미리 넣어 둔다(목록에 없으면 사용되지 않는다).
GLOBAL_BRAND_QUERY = {
    # 색상어 + 록밴드 Deep Purple + Purple Rain/Purple Heart 등과 충돌 → 정확 구문만
    "Purple": '"Purple mattress" OR "Purple Innovation" OR "Purple Mattress Company"',
    # 유령 캐릭터 Casper 와 충돌
    "Casper": '"Casper mattress" OR "Casper Sleep" OR "Casper bed"',
    # 일반 명사(꿀·과즙)와 충돌
    "Nectar": '"Nectar mattress" OR "Nectar Sleep" OR "Nectar bed"',
    # 비교적 고유하지만 오탐 여지가 있어 매트리스 맥락을 함께 요구
    "Serta": '"Serta mattress" OR "Serta Simmons" OR "Serta iComfort" OR "Serta Perfect Sleeper"',
    "Saatva": '"Saatva mattress" OR "Saatva bed"',
    "Sealy": '"Sealy mattress" OR "Sealy Posturepedic"',
}

# ── 국외 기사 관련성 필터 ────────────────────────────────────────────────
# 매트리스 맥락 단어. 하나라도 있으면 통과시킨다(음악 아티스트 협업 마케팅 등도 살린다).
_ON_TOPIC_WORDS = ("mattress", "mattresses", "bed", "beds", "bedding", "sleep", "sleeper",
                   "pillow", "pillows", "foam", "furniture")
# 명백히 무관한 분야(음악·연예) 매체
_OFF_TOPIC_SOURCES = ("loudwire", "rollingstone", "billboard", "pitchfork", "consequence",
                      "spin.com", "nme.com", "kerrang", "revolvermag", "ultimateclassicrock",
                      "brooklynvegan", "stereogum", "variety", "hollywoodreporter", "tmz")
# 밴드·앨범·공연 관련어
_OFF_TOPIC_WORDS = ("band", "album", "tour", "concert", "reunion", "rock", "guitarist",
                    "drummer", "bassist", "singer", "vocalist", "setlist", "gig",
                    "festival", "lineup", "song", "frontman", "discography", "tracklist")


def _strip_tags(html):
    """RSS description 의 HTML 태그 제거(관련성 판정용 평문)."""
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html or "")).strip()


def _has_word(text, words):
    """단어 경계 기준 포함 검사 — 첫 매칭 단어 반환(없으면 None).
       'rocket' 이 'rock' 으로 잡히는 식의 오탐을 막는다."""
    for w in words:
        if re.search(r"\b%s\b" % re.escape(w), text):
            return w
    return None


# 브랜드가 '매트리스 회사'임을 확정해 주는 표기 — 모호하지 않은 형태만 넣는다.
# ★ 'Purple' 단독은 색상·록밴드와 겹치므로 절대 넣지 않는다.
# 이 표기가 있으면 제목에 mattress/bed 가 없어도 통과시킨다. 그렇게 하지 않으면
# "Serta Simmons names new CEO", "Purple Innovation reports Q2 earnings" 같은
# 정작 중요한 경쟁사 실적·인사·M&A 기사가 전부 걸러진다.
GLOBAL_BRAND_ALIASES = {
    "Purple": ("purple innovation", "purple mattress"),
    "Serta": ("serta",),
    "Sleep Number": ("sleep number",),
    "Tempur-Pedic": ("tempur",),
    "Casper": ("casper sleep",),
    "Nectar": ("nectar sleep",),
    "Saatva": ("saatva",),
    "Sealy": ("sealy",),
}


def _brand_relevant(title, source, brand=None, desc=None):
    """국외(영문) 기사 관련성 판정 → (통과여부, 제외사유).
       판정 대상은 제목 + RSS 스니펫(description).
       1) 매트리스 관련어가 있으면 통과 (음악 아티스트 협업 마케팅 기사도 여기서 살아난다)
       2) 브랜드의 '모호하지 않은 회사 표기'가 있으면 통과 (실적·인사·M&A 기사 보호)
       3) 음악·연예 매체이거나 밴드 관련어가 있으면 제외
       4) 아무 근거도 없으면 제외"""
    t = (title or "").lower()
    body = (t + " " + _strip_tags((desc or "")).lower()).strip()
    s = (source or "").lower()
    if _has_word(body, _ON_TOPIC_WORDS):
        return True, None
    for alias in GLOBAL_BRAND_ALIASES.get(brand or "", ()):
        if alias in body:
            return True, None
    bad_src = next((d for d in _OFF_TOPIC_SOURCES if d in s), None)
    if bad_src:
        return False, "무관 매체(%s)" % bad_src
    bad_w = _has_word(t, _OFF_TOPIC_WORDS)
    if bad_w:
        return False, "밴드·공연 키워드(%s)" % bad_w
    return False, "매트리스 관련어 없음"

# 브랜드 기사 이미지 추출 동시 요청 수. 차단되면 낮출 것(보수적으로 4에서 시작).
BRAND_IMG_WORKERS = 4

# ── 신제품 동향 판정 (국내·국외 공통) ───────────────────────────────────
# 제외: 실적·재무·M&A·인사 기사. 섹션 제목이 '신제품·브랜드 동향'인데 이런 기사가 섞였다.
# ★ 실적은 '국내외 경쟁사 분기 실적' 섹션이 따로 있어 중복이다.
# ★ 'names' 단독은 넣지 않는다 — "Names New Collection" 같은 신제품 제목까지 잡는다.
#   인사 기사는 'names new ceo'/'appoints' 로 좁혀서 잡는다.
_NEWS_EXCLUDE = (
    # 실적·재무
    "reports q1", "reports q2", "reports q3", "reports q4",
    "first quarter", "second quarter", "third quarter", "fourth quarter",
    "quarter results", "quarterly results", "earnings", "revenue", "net sales",
    "fiscal year", "financial results", "guidance", "stock", "shares", "investor",
    "실적", "매출", "영업이익", "순이익", "주가", "공시",
    # 인수합병·조직
    "acquire", "acquisition", "merger", "buyout", "stake", "ipo",
    "employee count", "headcount", "layoff", "layoffs",
    "ceo", "appoints", "names new ceo", "steps down",
    "인수", "합병", "지분", "상장", "인사", "임명", "구조조정",
    # 소송·파산 — 명세 목록에는 없었으나 실제 수집에서 상위 카드에 올라왔다.
    # ("Serta Simmons: The Bill Comes Due. Court Awards $400 Million",
    #  "Rabbi Trust Funds as Property of the Estate: The Sleep Number...")
    # 실적·M&A 와 같은 성격(신제품 동향 아님)이라 함께 제외한다.
    "lawsuit", "court", "litigation", "verdict", "settlement", "bankruptcy",
    "chapter 11", "sues", "sued", "damages", "antitrust", "ruling", "trust funds",
    "소송", "판결", "법원", "파산", "과징금", "제재",
)
# 실적·주식·법률 전문 매체 — 제목만으로 놓칠 때를 위한 보조 신호
_NEWS_EXCLUDE_SOURCES = (
    "chapter11cases", "natlawreview", "law360", "lexology",
    "stocktitan", "tradingview", "quiverquant", "seekingalpha", "fool.com",
    "benzinga", "marketbeat", "simplywall", "investing.com", "zacks",
    "revelio", "yahoo finance",
)
# 우선: 신제품 발표로 보이는 신호
_NEWS_PRIORITY = (
    "launch", "launches", "unveil", "unveils", "introduce", "introduces",
    "debut", "debuts", "new collection", "new line", "releases", "announces new",
    "collaboration", "partnership",
    "출시", "론칭", "선봬", "선보", "공개", "신제품", "컬렉션", "협업",
)


def _find_kw(text, words):
    """제목에서 첫 매칭 키워드 반환(없으면 None).
       영문은 단어 경계, 한글은 부분 일치(조사가 붙기 때문)."""
    for w in words:
        if re.search(r"[a-z]", w):
            if re.search(r"\b%s\b" % re.escape(w), text):
                return w
        elif w in text:
            return w
    return None


def _news_grade(title, source=None):
    """신제품 관련성 등급 → ('exclude'|'priority'|'neutral', 걸린 키워드).
       1) 제외 키워드(또는 실적·법률 전문 매체) → 무조건 제외
       2) 우선 키워드 → 상위 노출
       3) 둘 다 없으면 중립 (리뷰·프로모션 기사가 여기 남는다)"""
    t = (title or "").lower()
    kw = _find_kw(t, _NEWS_EXCLUDE)
    if kw:
        return "exclude", kw
    s = (source or "").lower()
    bad_src = next((d for d in _NEWS_EXCLUDE_SOURCES if d in s), None)
    if bad_src:
        return "exclude", "매체:%s" % bad_src
    kw = _find_kw(t, _NEWS_PRIORITY)
    if kw:
        return "priority", kw
    return "neutral", None


# 브랜드 공식 도메인 → 무료 로고 서비스(Clearbit, 키 불필요)로 로고 URL 생성.
# 기사 사진이 없을 때 프런트에서 이 주소를 img src 로 바로 사용한다.
BRAND_DOMAINS = {
    "에이스침대": "acebed.com",
    "씰리": "sealy.co.kr",
    "한샘": "hanssem.com",
    "이케아": "ikea.com",
    "씰리침대": "sealy.co.kr",
    "슬럼버랜드": "slumberland.co.kr",
    "킹스다운": "kingsdown.co.kr",
    "지누스": "zinus.com",
    "Sleep Number": "sleepnumber.com",
    "Tempur-Pedic": "tempurpedic.com",
    "Purple": "purple.com",
    "Serta": "serta.com",
}


# 브랜드명 정규화(별칭·표기 차이 흡수) → BRAND_DOMAINS 정규 키. 공백 제거+소문자로 매칭.
_BRAND_CANON = {
    "에이스침대": "에이스침대", "ace": "에이스침대", "acebed": "에이스침대", "ace침대": "에이스침대",
    "이케아": "이케아", "ikea": "이케아", "이케아코리아": "이케아", "ikeakorea": "이케아",
    "한샘": "한샘", "hanssem": "한샘",
    "씰리": "씰리", "sealy": "씰리", "씰리침대": "씰리침대", "씰리코리아": "씰리침대",
    "슬럼버랜드": "슬럼버랜드", "slumberland": "슬럼버랜드",
    "킹스다운": "킹스다운", "kingsdown": "킹스다운",
    "지누스": "지누스", "zinus": "지누스",
    "sleepnumber": "Sleep Number",
    "tempur-pedic": "Tempur-Pedic", "tempurpedic": "Tempur-Pedic", "tempur": "Tempur-Pedic", "템퍼": "Tempur-Pedic",
    "purple": "Purple", "퍼플": "Purple",
    "serta": "Serta", "썰타": "Serta", "서타": "Serta",
}


def _brand_norm(name):
    """표기 차이 흡수: 공백 제거+소문자 후 별칭 사전 매칭 → 정규 브랜드명(미등록이면 원문)."""
    k = re.sub(r"\s+", "", (name or "")).lower()
    return _BRAND_CANON.get(k, name)


def _brand_logos(brand):
    """공용 로고 소스: (clearbit_url, favicon_fallback, matched). 상단·하단 카드가 함께 사용."""
    dom = BRAND_DOMAINS.get(_brand_norm(brand))
    if not dom:
        return (None, None, False)
    return ("https://logo.clearbit.com/" + dom,
            "https://www.google.com/s2/favicons?sz=128&domain=" + dom, True)


# (구 _OG_IMAGE_RE 제거 — 브랜드 이미지는 _IMG_METAS(og:image/twitter:image/image_src)를 쓴다)

# Google News 링크는 자체 뷰어의 일반 로고(lh3.googleusercontent 등)만 노출하므로,
# 아래 도메인의 이미지는 "대표 이미지 아님"으로 보고 버린다(→ 프런트에서 브랜드 이니셜로 대체).
_GENERIC_IMG_HOSTS = ("googleusercontent.com", "gstatic.com", "google.com")


# 매체 로고·기본 썸네일로 흔히 쓰이는 파일명 조각 → 대표 이미지로 쓰지 않는다.
_BAD_IMG_WORDS = ("logo", "default", "noimage", "no-image", "no_image",
                  "placeholder", "blank", "sprite", "avatar")


def _img_usable(img, base=None):
    """상대경로면 절대경로로 보정하고, 일반 로고/기본 썸네일이면 버린다.
       쓸 만하면 URL, 아니면 None."""
    if not img:
        return None
    img = img.strip()
    if base:
        img = urllib.parse.urljoin(base, img)
    if not img.startswith(("http://", "https://")):
        return None
    p = urllib.parse.urlparse(img)
    if any(h in p.netloc.lower() for h in _GENERIC_IMG_HOSTS):
        return None                       # 구글 뷰어 일반 로고
    name = (p.path.rsplit("/", 1)[-1] or "").lower()
    if any(w in name for w in _BAD_IMG_WORDS):
        return None                       # logo.png / default_thumb.jpg 등
    return img


def _meta_image(page_url, timeout=5):
    """기사 페이지의 '메타 태그만' 보고 대표 이미지를 뽑는다. (실패 사유와 함께 반환)
       ★ 본문 <img> 크롤링은 하지 않는다 — 매체마다 구조가 달라 광고가 잡힌다."""
    if not page_url or not page_url.startswith("http"):
        return None, "링크 없음"
    try:
        r = requests.get(page_url, headers=_GNEWS_UA, timeout=timeout, allow_redirects=True)
        r.raise_for_status()
        html = r.text
    except requests.exceptions.HTTPError as e:            # 403/404 등
        return None, "HTTP %s" % getattr(e.response, "status_code", "?")
    except requests.exceptions.Timeout:
        return None, "타임아웃(%ds)" % timeout
    except Exception as e:  # noqa: BLE001
        return None, "요청 실패(%s)" % type(e).__name__
    for rx in _IMG_METAS:                                  # og:image → twitter:image → image_src
        m = rx.search(html)
        if m:
            u = _img_usable(m.group(1), page_url)
            if u:
                return u, None
    return None, "메타 태그 없음/부적합"


def _brand_image(link, rss_media):
    """브랜드 기사 대표 이미지. (url, 획득경로|실패사유) 반환.
       1) RSS 이미지(media:content/thumbnail/enclosure)
       2) 기사 원문 메타 태그 — ★Google News 링크는 실제 기사 URL로 디코드한 뒤 조회
          (디코드 없이 열면 구글 뷰어 페이지의 일반 로고만 나와 항상 실패했다)
       3) 실패 → None (프런트에서 브랜드 로고로 폴백)"""
    u = _img_usable(rss_media)
    if u:
        return u, "rss"
    real = _decode_gnews_url(link) or link
    u, why = _meta_image(real)
    if u:
        return u, "meta"
    return None, why or "이미지 없음"


# RSS <item> 안의 이미지 태그(media:content / media:thumbnail / enclosure)
_MRSS = "{http://search.yahoo.com/mrss/}"


def _rss_media(item):
    """RSS item 의 media:content/thumbnail 또는 이미지 enclosure URL. 없으면 None."""
    try:
        for tag in (_MRSS + "content", _MRSS + "thumbnail"):
            el = item.find(tag)
            if el is not None and el.get("url", "").startswith("http"):
                return el.get("url")
        enc = item.find("enclosure")
        if enc is not None and enc.get("url", "").startswith("http") \
                and "image" in (enc.get("type") or "image"):
            return enc.get("url")
    except Exception:  # noqa: BLE001
        pass
    return None


def _product_name(title, brand, source):
    """기사 제목에서 상품명 추출: 따옴표 안 텍스트 우선 → 없으면 '- 언론사' 제거 + 앞 '브랜드,' 제거."""
    t = (title or "").strip()
    # 1) 따옴표(' ' " " ‘ ’ “ ”) 안 텍스트
    m = re.search(r'[\'"‘’“”]([^\'"‘’“”]{2,})[\'"‘’“”]', t)
    if m:
        return m.group(1).strip()
    # 2) 뒤쪽 " - 언론사" 제거
    if source and t.endswith(" - " + source):
        t = t[:-(len(source) + 3)]
    else:
        t = re.sub(r'\s*[-–—]\s*[^-–—]+$', '', t)
    # 3) 앞쪽 "브랜드," 제거
    if brand:
        t = re.sub(r'^\s*' + re.escape(brand) + r'\s*[,·:\-]?\s*', '', t)
    return t.strip() or (title or "").strip()


def _brand_news(brands, query_fn, hl, gl, featured_brands=None, items_brands=None):
    """브랜드별 신제품 뉴스를 모아 {status, items(≤6), featured(브랜드당 1건·≤3)} 반환.
       한 브랜드 검색이 실패해도 나머지는 반환. 전부 실패면 status=error. (국내·국외 공용)

       featured_brands / items_brands 를 주면 상단·하단이 서로 다른 브랜드를 본다.
       (수집은 brands 전체로 한 번만 하고 표시 단계에서 나눈다 — 중복 수집 없음)
       생략하면 종전처럼 양쪽 모두 전체 브랜드를 대상으로 한다."""
    lang = hl.split("-")[0]
    collected, media_by_link, brand_stat = [], {}, {}
    for brand in brands:
        try:
            url = ("https://news.google.com/rss/search?q=" + urllib.parse.quote(query_fn(brand))
                   + "&hl=%s&gl=%s&ceid=%s:%s" % (hl, gl, gl, lang))
            resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            root = ET.fromstring(resp.content)
        except Exception:  # noqa: BLE001 — 한 브랜드 실패해도 계속
            continue
        cnt, seen, dropped = 0, 0, []
        for item in root.iter("item"):
            title = (item.findtext("title") or "").strip()
            if not title:
                continue
            seen += 1
            link = (item.findtext("link") or "").strip()
            pub = item.findtext("pubDate") or ""
            src_el = item.find("source")
            source = (src_el.text.strip() if (src_el is not None and src_el.text) else "")
            # 국외(영문)만 관련성 필터. 검색어를 좁혀도 새어 나오는 무관 기사를 여기서 막는다.
            # ★ 필터는 append 전에 두어, 제외된 기사가 '브랜드당 3건' 슬롯을 먹지 않게 한다.
            if lang == "en":
                okrel, why = _brand_relevant(title, source, brand,
                                             item.findtext("description") or "")
                if not okrel:
                    dropped.append((title, source, why))
                    continue
            # 신제품 동향 판정(국내·국외 공통): 실적·M&A·인사 기사 제외
            grade, gkw = _news_grade(title, source)
            if grade == "exclude":
                dropped.append((title, source, "신제품 무관(%s)" % gkw))
                continue
            media = _rss_media(item)
            if link and media:
                media_by_link[link] = media
            collected.append({"brand": brand, "title": title, "source": source,
                              "date": _fmt_pubdate(pub), "link": link,
                              "grade": grade, "grade_kw": gkw,
                              "product_name": _product_name(title, brand, source)})
            cnt += 1
            if cnt >= 3:  # 브랜드별 최신 최대 3건(중복 제거 재료 확보)
                break
        brand_stat[brand] = cnt
        print("[brand_news][%s] 검색 %d건 → 통과 %d건 / 제외 %d건" % (brand, seen, cnt, len(dropped)))
        for t, s, why in dropped[:6]:
            print("    ✗ %s%s — %s" % (t[:64], (" (%s)" % s) if s else "", why))
        if cnt == 0:
            print("[brand_news][%s] ⚠ 필터 후 기사 0건 — 다른 브랜드 기사로 채웁니다" % brand)

    if not collected:
        return {"status": "error", "reason": "브랜드 뉴스를 받지 못했습니다"}

    collected.sort(key=lambda x: x["date"], reverse=True)  # 최신순

    # 이미지 확보(RSS media → 기사 메타 태그): 중복 판정(이미지 URL 신호)·대표 선정에 필요.
    # 요청 간 300ms 대기(기사 수만큼 요청이 나간다). 실패는 조용히 로고 폴백.
    # ★ 성능: 기사 1건당 요청 3회(Google News 디코드 GET+POST, 메타 GET)가 나가
    #   순차로 돌리면 국내 24건에 70초가 걸렸다(전체의 90%). 스레드 풀로 병렬화한다.
    #   기사마다 언론사가 달라 특정 호스트에 몰리지 않으므로 sleep 없이 동시 4개로 제한.
    img_cache, img_stat, img_fail = {}, {"rss": 0, "meta": 0}, []
    links = list(dict.fromkeys(it["link"] for it in collected if it.get("link")))
    if links:
        with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(BRAND_IMG_WORKERS, len(links))) as _ex:
            for lk, res in zip(links, _ex.map(
                    lambda l: _brand_image(l, media_by_link.get(l)), links)):
                img_cache[lk] = res
    for it in collected:
        img, how = img_cache.get(it.get("link"), (None, "링크 없음"))
        it["image"] = img
        if img:
            img_stat[how] = img_stat.get(how, 0) + 1
        else:
            img_fail.append((it["brand"], it["title"][:40], it["link"], how))
    print("[brand_news] 이미지 확보 %d건(RSS %d · 메타 %d) / 로고 폴백 %d건"
          % (img_stat["rss"] + img_stat["meta"], img_stat["rss"], img_stat["meta"], len(img_fail)))
    for b, t, lk, why in img_fail:
        print("  · 폴백 [%s] %s — %s | %s" % (b, t, why, lk[:70]))

    # 로고: 상단·하단 공용 소스(_brand_logos)로 모든 기사에 부여 + 매칭 성공/실패 로그
    logo_ok, logo_fail = set(), set()
    for it in collected:
        lu, lf, matched = _brand_logos(it["brand"])
        it["logo_url"], it["logo_fallback"] = lu, lf
        (logo_ok if matched else logo_fail).add(it["brand"])
    print("[brand_news] 로고 매칭 성공: %s" % (", ".join(sorted(logo_ok)) or "-"))
    if logo_fail:
        print("[brand_news] 로고 매칭 실패(→첫글자 폴백): %s" % ", ".join(sorted(logo_fail)))

    # 중복 제거 — 시몬스 소식과 동일 공용 로직, 단 같은 브랜드 안에서만 묶는다(★타 브랜드 병합 금지)
    groups = _cluster_articles(collected, same_fn=lambda a, b: a["brand"] == b["brand"])
    multi = [g for g in groups if len(g["items"]) > 1]
    print("[brand_news] 중복 그룹 %d개(2건+):" % len(multi))
    for gi, g in enumerate(multi, 1):
        rep = sorted(g["items"], key=_rep_key)[0]
        print("  [그룹 %d][%s] %d건 — 대표: %s" % (gi, rep["brand"], len(g["items"]), rep["title"]))
        how = {t: (r, s) for (t, r, s) in g["merges"]}
        for m in g["items"]:
            rs = how.get(m["title"])
            via = ("  ← %s %.2f" % (rs[0], rs[1])) if rs else "  (기준)"
            print("     · %s%s%s" % (m["title"], via, "   ★대표" if m is rep else ""))

    reps = [sorted(g["items"], key=_rep_key)[0] for g in groups]  # 그룹 대표: 이미지>날짜>길이
    reps.sort(key=lambda x: x["date"], reverse=True)
    # 정렬 우선순위: ① 우선 등급 → ② 기사 사진 있는 것 → ③ 최신순
    # (파이썬 sort 는 안정 정렬이라, 위에서 맞춰 둔 날짜 순서가 동점 구간에 그대로 남는다)
    # ★ ②는 보조 신호일 뿐이다 — 로고 폴백이라고 제외하지는 않는다(og:image 추출 실패일 수도 있다).
    reps.sort(key=lambda x: (0 if x.get("grade") == "priority" else 1,
                             0 if x.get("image") else 1))

    def _out(it):
        return {"brand": it["brand"], "title": it["title"], "source": it["source"],
                "date": it["date"], "link": it["link"], "image": it.get("image"),
                "product_name": it.get("product_name"),
                "logo_url": it.get("logo_url"), "logo_fallback": it.get("logo_fallback")}

    # 상단·하단이 볼 기사 풀을 나눈다(목록을 주지 않으면 종전대로 전체).
    feat_pool = [r for r in reps if r["brand"] in featured_brands] if featured_brands else reps
    item_pool = [r for r in reps if r["brand"] in items_brands] if items_brands else reps
    if featured_brands or items_brands:
        print("[brand_news] 상단 대상 %s → %d건 / 하단 대상 %s → %d건"
              % (featured_brands or "전체", len(feat_pool), items_brands or "전체", len(item_pool)))

    # featured(대표 출시 상품): 브랜드당 1건 → 우선 등급·사진 순으로 3건
    best = {}
    for r in feat_pool:                  # reps 는 이미 (우선→중립, 사진, 최신) 순
        best.setdefault(r["brand"], r)
    featured = list(best.values())[:3]
    # 카드가 3칸이므로 항상 3개를 채운다 — 브랜드가 모자라면 중립 기사까지 끌어 쓴다.
    if len(featured) < 3:
        have = {id(f) for f in featured}
        for r in feat_pool:
            if len(featured) >= 3:
                break
            if id(r) not in have:
                featured.append(r)
                have.add(id(r))
        print("[brand_news] featured 부족 → 다른 기사로 보충해 %d건" % len(featured))

    # items: 브랜드별 최소 1건 보장 → 나머지(우선→중립 순)로 채워 6건.
    # ★ 상단에 쓴 기사는 하단에서 뺀다 — 같은 기사가 두 번 보이지 않게.
    #   (상·하단 브랜드가 겹치지 않으면 원래 안 겹치지만, 한 기사가 여러 브랜드
    #    쿼리에 동시에 걸릴 수 있어 링크 기준으로 한 번 더 막는다)
    used_links = {f.get("link") for f in featured if f.get("link")}
    pool = [r for r in item_pool if r.get("link") not in used_links]
    dropped_dup = len(item_pool) - len(pool)
    if dropped_dup:
        print("[brand_news] 상단과 중복된 기사 %d건을 하단에서 제외" % dropped_dup)
    items, picked = [], set()
    for r in pool:  # 브랜드별 대표 먼저(브랜드 누락 방지)
        if r["brand"] not in {x["brand"] for x in items}:
            items.append(r)
            picked.add(id(r))
    for r in pool:  # 남은 슬롯 채움
        if id(r) not in picked and len(items) < 6:
            items.append(r)
            picked.add(id(r))
    items = sorted(items[:6], key=lambda x: x["date"], reverse=True)
    # 그래도 모자라면 빈 칸 대신 안내 카드를 채운다.
    while len(featured) < 3:
        featured.append({"brand": "", "title": "최근 관련 기사 없음", "source": "",
                         "date": "", "link": "", "image": None, "grade": "empty",
                         "product_name": "최근 관련 기사 없음",
                         "logo_url": None, "logo_fallback": None})
        print("[brand_news] ⚠ 표시할 기사 부족 → '최근 관련 기사 없음' 카드 추가")

    empty = [b for b, c in brand_stat.items() if c == 0]
    if empty:
        print("[brand_news] ⚠ 필터 후 0건 브랜드: %s" % ", ".join(empty))
    print("[brand_news] 브랜드별 통과 건수: %s"
          % ", ".join("%s %d" % (b, c) for b, c in brand_stat.items()))
    print("[brand_news] 최종 선택 featured:")
    for f in featured:
        g = {"priority": "우선", "neutral": "중립", "empty": "빈칸"}.get(f.get("grade"), "중립")
        kw = (" ← %s" % f.get("grade_kw")) if f.get("grade_kw") else ""
        print("    ★ [%s] %s (%s · %s) [%s]%s"
              % (f["brand"], f["title"][:56], f["source"], f["date"], g, kw))
    print("[brand_news] 최종 표시 items %d / featured %d" % (len(items), len(featured)))
    return {"status": "ok", "items": [_out(x) for x in items],
            "featured": [_out(x) for x in featured]}


def _domestic_query(brand):
    """이름 충돌 브랜드는 DOMESTIC_BRAND_QUERY 의 정확 구문을, 나머지는 기본 쿼리를 쓴다.
       (검색어는 _brand_news 안에서 urllib.parse.quote 로 인코딩된다)"""
    q = DOMESTIC_BRAND_QUERY.get(brand)
    return q if q else '"%s" (신제품 OR 출시 OR 론칭)' % brand


def update_domestic():
    """국내 브랜드 신제품 뉴스 (한국어).
       상단: 한샘·에이스침대·이케아 / 하단: 씰리침대·템퍼·슬럼버랜드·킹스다운·지누스."""
    return _brand_news(DOMESTIC_BRANDS, _domestic_query, "ko", "KR",
                       featured_brands=DOMESTIC_FEATURED_BRANDS,
                       items_brands=DOMESTIC_ITEM_BRANDS)


def _global_query(brand):
    """이름 충돌 브랜드는 GLOBAL_BRAND_QUERY 의 정확 구문을, 나머지는 기본 쿼리를 쓴다."""
    q = GLOBAL_BRAND_QUERY.get(brand)
    return q if q else '"%s" (new OR launch OR release OR mattress)' % brand


def update_global_brands():
    """국외 브랜드(Sleep Number·Tempur-Pedic·Purple·Serta) 신제품 뉴스 (영어)."""
    return _brand_news(GLOBAL_BRANDS, _global_query, "en-US", "US")


# ── 경쟁사 분석 — SEC EDGAR (무료·무키, 미국 상장사) ─────────────────────
# SEC 규칙: data.sec.gov / www.sec.gov 요청에는 반드시 User-Agent 헤더 필요(없으면 403)
SEC_HEADERS = {"User-Agent": "Simmons Dashboard contact@example.com"}
# 국외(Global) 경쟁사: (표시명, 티커, 폴백 CIK, 로고 도메인). CIK 매칭 실패 시 폴백 사용.
# 참고: SEC 파일상 Sleep Number 티커는 'SNBRQ', Tempur Sealy는 사명 변경(Somnigroup, CIK 1206264).
GLOBAL_COMPETITORS = [
    ("Sleep Number", "SNBR", 827187, "sleepnumber.com"),
    ("Tempur Sealy", "TPX", 1206264, "tempursealy.com"),
    ("Purple Innovation", "PRPL", 1643953, "purple.com"),
    ("Leggett & Platt", "LEG", 58492, "leggett.com"),
]
REVENUE_KEYS = [
    "RevenueFromContractWithCustomerExcludingAssessedTax",
    "Revenues",
    "SalesRevenueNet",
]
_cik_cache = None


def _sec_get(url):
    resp = requests.get(url, timeout=30, headers=SEC_HEADERS)
    resp.raise_for_status()
    return resp.json()


def _load_cik_map():
    """티커(대문자) → CIK(int). company_tickers.json 1회 로드 후 캐시."""
    global _cik_cache
    if _cik_cache is not None:
        return _cik_cache
    data = _sec_get("https://www.sec.gov/files/company_tickers.json")
    m = {}
    for v in data.values():
        try:
            m[str(v.get("ticker", "")).upper()] = int(v.get("cik_str"))
        except Exception:  # noqa: BLE001
            pass
    _cik_cache = m
    return m


def _resolve_cik(ticker, cikmap, fallback):
    """티커 → CIK. 정확 매칭 → 접두 매칭(SNBR→SNBRQ) → 폴백."""
    t = str(ticker).upper()
    if t in cikmap:
        return cikmap[t]
    for k, v in cikmap.items():
        if k.startswith(t):
            return v
    return fallback


def _quarterly_series(facts, keys, unit="USD"):
    """분기(약 3개월, 10-Q) 값. return {(fy:int, fp:str): {'val','end','label'}}.
       10-Q 이고 기간이 ~90일(단일 분기)인 항목만 사용해 YTD(누적) 값을 배제한다."""
    usgaap = facts.get("facts", {}).get("us-gaap", {})
    for k in keys:
        node = usgaap.get(k)
        if not node:
            continue
        arr = node.get("units", {}).get(unit)
        if not arr:
            continue
        out = {}
        for d in arr:
            val, fp, form = d.get("val"), d.get("fp"), d.get("form")
            start, end, fy = d.get("start"), d.get("end"), d.get("fy")
            if val is None or not start or not end or form != "10-Q":
                continue
            if fp not in ("Q1", "Q2", "Q3", "Q4"):
                continue
            try:
                s = datetime.datetime.strptime(start, "%Y-%m-%d").date()
                e = datetime.datetime.strptime(end, "%Y-%m-%d").date()
            except Exception:  # noqa: BLE001
                continue
            if not (80 <= (e - s).days <= 100):  # 단일 분기(3개월)만 → YTD 제외
                continue
            key = (int(fy) if fy is not None else e.year, fp)
            if key not in out or e > out[key]["end"]:  # 정정치는 end 최신 우선
                out[key] = {"val": val, "end": e, "label": "%s %s" % (key[0], fp)}
        if out:
            return out
    return {}


def _quarter_metric(facts, keys):
    """가장 최근 분기의 (값, 전년 동기 대비 %, 라벨). 없으면 (None, None, None)."""
    qs = _quarterly_series(facts, keys)
    if not qs:
        return (None, None, None)
    latest = max(qs, key=lambda kk: qs[kk]["end"])
    val = qs[latest]["val"]
    fy, fp = latest
    prev = qs.get((fy - 1, fp))  # 전년 동일 분기
    yoy = None
    if prev and prev["val"]:
        yoy = round((val - prev["val"]) / abs(prev["val"]) * 100.0, 1)
    return (val, yoy, qs[latest]["label"])


def _logo_candidates(domain):
    """로고 자동 소스 후보(앞에서부터 시도, 프런트에서 onerror 체인으로 대체):
       1) Clearbit  2) Google 파비콘(고해상)  3) 도메인 파비콘 직접."""
    return [
        "https://logo.clearbit.com/%s" % domain,
        "https://www.google.com/s2/favicons?domain=%s&sz=128" % domain,
        "https://%s/favicon.ico" % domain,
    ]


def _usd_krw_rate():
    """경쟁사 금액(USD) → 원화 환산용 USD/KRW 환율. 실패 시 None(원화 표시 생략)."""
    try:
        r = requests.get("https://api.frankfurter.app/latest?from=USD&to=KRW",
                         timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        return r.json()["rates"]["KRW"]
    except Exception:  # noqa: BLE001
        return None


# ── 국내 경쟁사 — 네이버 금융 크롤링 (무키) ──────────────────────────────
# 종목: 에이스침대 003800, 지누스 013890, 한샘 009240. (도메인은 로고용)
NAVER_COMPANIES = [
    ("에이스침대", "003800", "acebed.com"),
    ("지누스", "013890", "zinus.co.kr"),
    ("한샘", "009240", "hanssem.com"),
    ("현대리바트", "079430", "hyundailivart.co.kr"),
    ("코웨이", "021240", "coway.co.kr"),
    ("에넥스", "011090", "enex.co.kr"),
]


def _naver_financials(code):
    """네이버 금융 종목 페이지의 '기업실적분석' 표에서 최근 분기 매출/순이익(원화) + YoY.
       금액 단위는 억원 → 원(×1e8)으로 환산. 실패/구조 변경 시 None."""
    from bs4 import BeautifulSoup  # 지연 임포트 — 미설치여도 서버는 계속 동작

    url = "https://finance.naver.com/item/main.naver?code=%s" % code
    resp = requests.get(url, timeout=20,
                        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or "utf-8"  # 네이버 금융은 현재 UTF-8
    soup = BeautifulSoup(resp.text, "html.parser")
    area = soup.select_one("div.cop_analysis")
    if not area:
        return None
    table = area.select_one("table")
    if not table:
        return None

    trs = table.select("thead tr")
    # 분기 컬럼 수 (상단 헤더 '분기' colspan) — 보통 연간4 + 분기6
    nq = 4
    if trs:
        for th in trs[0].find_all("th"):
            if "분기" in th.get_text():
                try:
                    nq = int(th.get("colspan") or 4)
                except Exception:  # noqa: BLE001
                    nq = 4
    period_ths = trs[1].find_all("th") if len(trs) > 1 else []
    periods = [th.get_text(strip=True) for th in period_ths
               if re.match(r"\d{4}\.\d{2}", th.get_text(strip=True))]

    def row_vals(metric, idx):
        """지표명(우선) 또는 tbody 행 위치(폴백)로 값 목록 반환."""
        rows = table.select("tbody tr")
        for tr in rows:
            th = tr.find("th")
            if th and metric in th.get_text():
                return [td.get_text(strip=True) for td in tr.find_all("td")]
        if idx < len(rows):
            return [td.get_text(strip=True) for td in rows[idx].find_all("td")]
        return None

    rev = row_vals("매출액", 0)
    ni = row_vals("당기순이익", 2)
    if not rev or not ni or not periods:
        return None

    q_periods = periods[-nq:]

    def _qslice(v):
        return v[-nq:] if len(v) >= nq else v

    rev_q, ni_q = _qslice(rev), _qslice(ni)

    def num(s):
        s = (s or "").replace(",", "").replace("+", "").strip()
        if s in ("", "-", "N/A"):
            return None
        try:
            return float(s)
        except Exception:  # noqa: BLE001
            return None

    def pmap(vals):
        mp = {}
        for p, val in zip(q_periods, vals):
            m = re.match(r"(\d{4})\.(\d{2})", p)
            if m:
                x = num(val)
                if x is not None:
                    mp[(m.group(1), m.group(2))] = x
        return mp

    rev_map, ni_map = pmap(rev_q), pmap(ni_q)

    # 최신 실제 분기(추정치 '(E)' 제외, 값 존재)
    latest = None
    for i in range(len(q_periods) - 1, -1, -1):
        if "(E)" in q_periods[i]:
            continue
        if i < len(rev_q) and num(rev_q[i]) is not None:
            latest = i
            break
    if latest is None:
        return None
    mm = re.match(r"(\d{4})\.(\d{2})", q_periods[latest])
    y, mo = mm.group(1), mm.group(2)

    def yoy(mp):  # 전년 동기 대비
        cur, prev = mp.get((y, mo)), mp.get((str(int(y) - 1), mo))
        if cur is not None and prev not in (None, 0):
            return round((cur - prev) / abs(prev) * 100.0, 1)
        return None

    rev_v = num(rev_q[latest])
    ni_v = num(ni_q[latest]) if latest < len(ni_q) else None
    qmap = {"03": "Q1", "06": "Q2", "09": "Q3", "12": "Q4"}
    return {
        "quarter": "%s %s" % (y, qmap.get(mo, mo)),
        "revenue_krw": (rev_v * 1e8) if rev_v is not None else None,      # 억원 → 원
        "revenue_yoy": yoy(rev_map),
        "net_income_krw": (ni_v * 1e8) if ni_v is not None else None,
        "net_income_yoy": yoy(ni_map),
    }


def update_domestic_financials():
    """네이버 금융 크롤링으로 국내 3사 최근 분기 매출·순이익(원화)+YoY.
       회사별 실패는 건너뛰고, 전부 실패면 {'status':'데이터 없음'}."""
    out = []
    for name, code, domain in NAVER_COMPANIES:
        try:
            fin = _naver_financials(code)
        except Exception:  # noqa: BLE001 — 절대 죽지 않게
            fin = None
        if not fin or (fin.get("revenue_krw") is None and fin.get("net_income_krw") is None):
            continue
        entry = {"name": name, "code": code, "logo_urls": _logo_candidates(domain)}
        entry.update(fin)
        out.append(entry)
    if not out:
        return {"status": "데이터 없음"}
    return out


def update_competitors():
    """경쟁사 분석 — 국외(Global) 4곳의 최근 분기(10-Q) 매출·순이익 + 전년 동기 대비(YoY).
       국내(Korea)는 네이버 금융 크롤링. 한 회사 실패해도 나머지 반환."""
    try:
        cikmap = _load_cik_map()
    except Exception:  # noqa: BLE001 — 티커 파일 실패해도 폴백 CIK로 진행
        cikmap = {}
    usd_krw = _usd_krw_rate()  # USD→KRW 환산 환율(없으면 None)

    glob, any_ok = [], False
    for name, ticker, fallback, domain in GLOBAL_COMPETITORS:
        entry = {"name": name, "ticker": ticker, "quarter": None,
                 "revenue": None, "revenue_yoy": None,
                 "net_income": None, "net_income_yoy": None,
                 "logo_urls": _logo_candidates(domain)}
        cik = _resolve_cik(ticker, cikmap, fallback)
        if cik is None:
            entry["error"] = "CIK 없음"
            glob.append(entry)
            continue
        try:
            facts = _sec_get("https://data.sec.gov/api/xbrl/companyfacts/CIK%s.json"
                             % str(cik).zfill(10))
            qrev, qrev_yoy, qlabel = _quarter_metric(facts, REVENUE_KEYS)
            qni, qni_yoy, qlabel2 = _quarter_metric(facts, ["NetIncomeLoss"])
            entry["quarter"] = qlabel or qlabel2
            entry["revenue"], entry["revenue_yoy"] = qrev, qrev_yoy
            entry["net_income"], entry["net_income_yoy"] = qni, qni_yoy
            if qrev is not None or qni is not None:
                any_ok = True
        except Exception as e:  # noqa: BLE001
            entry["error"] = str(e)
        glob.append(entry)

    try:
        korea = update_domestic_financials()
    except Exception:  # noqa: BLE001
        korea = {"status": "데이터 없음"}
    if not any_ok:
        return {"status": "error", "reason": "SEC 분기 데이터를 받지 못했습니다",
                "global": glob, "korea": korea, "usd_krw_rate": usd_krw}
    return {"status": "ok", "global": glob, "korea": korea, "usd_krw_rate": usd_krw}


# ── 환율 EUR/KRW — Frankfurter(ECB 기반, 무료·무키) ──────────────────────
def update_fx():
    """Frankfurter(ECB 기반, 무키·무료)로 USD/EUR/JPY 대비 원화(KRW) 환율.
       - 최근 12개월(+여유) 시계열을 통화별로 받아 현재값·전일대비(직전 영업일)·추이를 구성.
       - JPY는 100엔 단위(×100)로 환산해 표시한다.
       - rows: 상단 시세표(현재기준율/전일대비/전일기준율), series: 그래프용 시계열.
       실패 시 status=error, 서버는 죽지 않음."""
    base = "https://api.frankfurter.app"
    hdr = {"User-Agent": "Mozilla/5.0"}
    curs = ["USD", "EUR", "JPY"]
    try:
        today = datetime.date.today()
        start = (today - datetime.timedelta(days=400)).isoformat()  # 12개월 + 여유
        end = today.isoformat()
        raw = {}  # cur -> {date: KRW값}
        for c in curs:
            resp = requests.get("%s/%s..%s?from=%s&to=KRW" % (base, start, end, c),
                                timeout=25, headers=hdr)
            resp.raise_for_status()
            rr = resp.json().get("rates", {})
            raw[c] = {d: obj["KRW"] for d, obj in rr.items() if obj.get("KRW") is not None}
        # 세 통화가 공유하는 ECB 영업일 전체(정렬)
        dates = sorted(set().union(*[set(raw[c].keys()) for c in curs]))
        if not dates:
            return {"status": "error", "reason": "환율 시계열 없음"}
        cutoff = (today - datetime.timedelta(days=366)).isoformat()  # 최근 12개월만 유지
        dates = [d for d in dates if d >= cutoff]

        def scale(cur, v):  # JPY는 100엔 단위
            return round(v * 100.0, 2) if cur == "JPY" else round(v, 2)

        series = {"dates": dates}
        rows = []
        for c in curs:
            vals = [scale(c, raw[c][d]) if d in raw[c] else None for d in dates]
            series[c] = vals
            nn = [v for v in vals if v is not None]
            now = nn[-1] if nn else None
            prev = nn[-2] if len(nn) >= 2 else None
            change = round(now - prev, 2) if (now is not None and prev is not None) else None
            rows.append({"cur": c, "now": now, "prev": prev, "change": change})
        return {"status": "ok", "rows": rows, "series": series,
                "source": "Frankfurter (ECB 기반)"}
    except requests.exceptions.RequestException as e:  # noqa: BLE001
        return {"status": "error", "reason": "환율 조회 실패: %s" % e}
    except Exception as e:  # noqa: BLE001
        return {"status": "error", "reason": "환율 처리 실패: %s" % e}


# 섹션별 fetcher — 요청 한 번에 모두 실행. 추후 같은 패턴으로 확장 가능.
def update_usd_krw():
    """원자재(USD/톤) 원화 환산용 USD/KRW 환율. 실패 시 status=error."""
    r = _usd_krw_rate()
    if r is None:
        return {"status": "error", "rate": None}
    return {"status": "ok", "rate": r}


SR_XLSX_URL = ("https://www.sea-intelligence.com/images/press_docs/GLP-May2026/"
               "20260527_-_Sea-Intelligence_GLP_Press_Release_-_May_2026.xlsx")
SR_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def update_schedule_reliability():
    """★ 순수 추가: sea_intelligence 로 최신 GLP 편을 찾아 쓰고, 실패하면 아래 기존
       고정 URL 로직으로 폴백한다(하드코딩 URL 은 2026-04 에서 멈춰 있었다).

       Sea-Intelligence 글로벌 해상 정시성(Global Schedule Reliability) 월별 시계열.
       xlsx의 'Fig 1' 시트(행=연도 2021~2026, 열=Jan~Dec, 값=정시성 비율)를 파싱한다.
       접근 차단(403)·타임아웃·형식 오류 등은 서버가 죽지 않도록 status=error + reason 처리."""
    if update_schedule_reliability_auto is not None:
        res = update_schedule_reliability_auto(verbose=True)
        if res.get("status") == "ok":
            return res
        print("[sr] 자동 탐색 실패 → 기존 고정 URL 로 폴백: %s" % res.get("reason"))
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/125.0.0.0 Safari/537.36"}
    try:
        import io
        import openpyxl
        r = requests.get(SR_XLSX_URL, headers=hdr, allow_redirects=True, timeout=15)
        r.raise_for_status()
        if r.content[:2] != b"PK":  # xlsx=zip. 아니면 차단 페이지 등
            return {"status": "error", "reason": "엑셀이 아닌 응답(사이트 접근 차단 가능)"}
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        if "Fig 1" not in wb.sheetnames:
            return {"status": "error", "reason": "'Fig 1' 시트를 찾을 수 없음"}
        ws = wb["Fig 1"]
        # 1행 = [None, Jan, Feb, ... Dec] : 월 헤더로 12개월 열 위치 확정
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        month_cols = [i for i, v in enumerate(header)
                      if isinstance(v, str) and v.strip()[:3] in SR_MONTHS]
        if len(month_cols) < 12:  # 헤더 파싱 실패 시 관례 위치(B~M)로 대체
            month_cols = list(range(1, 13))
        years = {}
        for row in ws.iter_rows(min_row=2):
            cells = [c.value for c in row]
            if not cells:
                continue
            try:
                yi = int(cells[0])
            except (TypeError, ValueError):
                continue
            if not (2000 <= yi <= 2100):  # 연도 행만 채택
                continue
            vals = []
            for ci in month_cols:
                v = cells[ci] if ci < len(cells) else None
                if isinstance(v, (int, float)):
                    pct = v * 100.0 if v <= 1.5 else float(v)  # 비율→% 환산
                    vals.append(round(pct, 1))
                else:
                    vals.append(None)
            years[str(yi)] = vals
        if not years:
            return {"status": "error", "reason": "연도별 데이터 없음"}
        return {"status": "ok", "months": SR_MONTHS, "years": years,
                "source": "Sea-Intelligence"}
    except requests.exceptions.RequestException as e:  # noqa: BLE001
        return {"status": "error", "reason": "다운로드 실패: %s" % e}
    except Exception as e:  # noqa: BLE001
        return {"status": "error", "reason": "파싱 실패: %s" % e}


# ── 국제유가 — 한국석유공사 PETRONET 일일국제제품가격(월별, 3개 유종) ──────────
OIL_GET_URL = "https://www.petronet.co.kr/v4/excel/KDFQ0200_x2.jsp"
OIL_POST_URL = "https://www.petronet.co.kr/v4/sub.jsp"
# (제품코드, 응답키, 표시라벨) — 요청 ProdCDList 순서 = 응답 표의 열 순서
OIL_PRODS = [
    ("B001", "gasoline95", "휘발유(95RON)"),
    ("D008", "diesel005", "경유(0.05%)"),
    ("F001", "naphtha", "나프타"),
]


def _oil_num(cell):
    v = re.sub(r"<[^>]+>", "", cell).replace("&nbsp;", " ").strip()
    try:
        return round(float(v), 2)
    except ValueError:
        return None  # '-' 등 결측 → None


def _oil_parse(text):
    """월별 제품가 HTML 표 → rows[{period:'YYYY-MM', <9키>}].
       - 날짜 셀에 연도 없음. 다만 1월 행은 'YY년 01월'로 연도 표기 → 연도 확정.
         그 외 'MM월' 행은 직전 연도 유지(월 감소 시 +1 롤오버 안전장치).
       - 첫 셀이 월 형태가 아닌 행(헤더/전월비/전년동월비/평균)은 제외."""
    n_prod = len(OIL_PRODS)
    rows = []
    year, prev_m = None, None
    for tr in re.findall(r"<tr.*?>(.*?)</tr>", text, re.S | re.I):
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, re.S | re.I)
        if len(cells) < 1 + n_prod:
            continue
        first = re.sub(r"<[^>]+>", "", cells[0]).replace("&nbsp;", " ").strip()
        mfull = re.match(r"^\s*(\d{2})\s*년\s*(\d{1,2})\s*월", first)  # 'YY년 MM월'
        mmon = re.match(r"^\s*(\d{1,2})\s*월\s*$", first)             # 'MM월'
        if mfull:
            year = 2000 + int(mfull.group(1))
            mo = int(mfull.group(2))
        elif mmon:
            mo = int(mmon.group(1))
            if year is None:
                year = 2011  # FromDate 시작연도
            elif prev_m is not None and mo < prev_m:
                year += 1
        else:
            continue  # 헤더·요약행 제외
        prev_m = mo
        vals = [_oil_num(c) for c in cells[1:1 + n_prod]]
        row = {"period": "%04d-%02d" % (year, mo)}
        for (_, key, _), v in zip(OIL_PRODS, vals):
            row[key] = v
        rows.append(row)
    return rows


def update_oil_prices():
    """PETRONET 일일국제제품가격(월별, USD/배럴, 3개 유종, 2011.07~현재).
       GET(excel jsp) 먼저 시도 → 리다이렉트/빈 응답이면 POST(sub.jsp)로 폴백.
       실패 시 status=error, 서버는 죽지 않음."""
    ua = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
          "(KHTML, like Gecko) Chrome/125.0 Safari/537.36")
    ref = "https://www.petronet.co.kr/v4/sub.jsp"
    prods = [p[0] for p in OIL_PRODS]
    try:
        today = datetime.date.today()
        todate = "%04d%02d" % (today.year, today.month)
        rows = []
        # 1순위 GET
        try:
            qs = {"term": "m", "by": "2011", "bq": "3", "bm": "07", "bw": "1", "bd": "16",
                  "ay": str(today.year), "aq": str((today.month - 1) // 3 + 1),
                  "am": "%02d" % today.month, "aw": "1", "ad": "%02d" % today.day,
                  "ProdCDList": ",".join(prods)}
            g = requests.get(OIL_GET_URL, params=qs,
                             headers={"User-Agent": ua, "Referer": ref}, timeout=30)
            gt = g.content.decode("utf-8", "replace")
            if g.status_code == 200 and "<tr" in gt.lower():
                rows = _oil_parse(gt)
        except requests.exceptions.RequestException:
            rows = []
        # 2순위 POST 폴백
        if not rows:
            inner = "\\,".join("\\'%s\\'" % c for c in prods)  # \'B001\'\,\'D008\'\,\'F001\'
            param = ":T='M',:FromDate='201107',:ToDate='%s',:ProdCD='%s '" % (todate, inner)
            data = [("fmuId", "KDFQSTAT"), ("smuId", "KDFQ01"), ("tmuId", "KDFQ0200"),
                    ("fmuOrd", "03"), ("smuOrd", "03_01"), ("tmuOrd", "03_01_02"),
                    ("Parameter", param), ("ProdCDList", ",".join(prods)),
                    ("firstFlag", "T"), ("term", "m"),
                    ("by", "2011"), ("bq", "3"), ("bm", "07"), ("bw", "1"), ("bd", "16"),
                    ("ay", str(today.year)), ("aq", str((today.month - 1) // 3 + 1)),
                    ("am", "%02d" % today.month), ("aw", "1"), ("ad", "%02d" % today.day)]
            data += [("ProdCd", p) for p in prods]  # ProdCd 반복
            pr = requests.post(OIL_POST_URL, data=data, timeout=40,
                               headers={"User-Agent": ua, "Referer": ref,
                                        "Content-Type": "application/x-www-form-urlencoded"})
            pr.raise_for_status()
            txt = pr.content.decode("utf-8", "replace")
            if txt.count("�") > 10:  # utf-8 깨지면 euc-kr 폴백
                txt = pr.content.decode("euc-kr", "replace")
            rows = _oil_parse(txt)
        if not rows:
            return {"status": "error", "reason": "유가 데이터 행 없음(형식 변경/차단 가능)"}
        return {"status": "ok", "source": "PETRONET 일일국제제품가격", "unit": "USD/배럴",
                "series": [{"key": k, "label": lbl} for _, k, lbl in OIL_PRODS], "rows": rows}
    except requests.exceptions.RequestException as e:  # noqa: BLE001
        return {"status": "error", "reason": "PETRONET 조회 실패: %s" % e}
    except Exception as e:  # noqa: BLE001
        return {"status": "error", "reason": "유가 파싱 실패: %s" % e}


FETCHERS = {
    "simmons_news": update_simmons_news,
    "usd_krw": update_usd_krw,
    "schedule_reliability": update_schedule_reliability,
    "oil_prices": update_oil_prices,
    "domestic": update_domestic,
    "global_brands": update_global_brands,
    "competitors": update_competitors,
    "fx": update_fx,
}
if compute_icis_forecast is not None:  # 순수 추가: 예측 재계산을 업데이트 버튼에 덧붙임
    FETCHERS["icis_forecast"] = compute_icis_forecast
if compute_sr_forecast is not None:  # 순수 추가: 해상 정시성 예측도 업데이트에 덧붙임
    FETCHERS["sr_forecast"] = compute_sr_forecast
if compute_oil_forecast is not None:  # 순수 추가: 국제유가 예측도 업데이트에 덧붙임
    FETCHERS["oil_forecast"] = compute_oil_forecast
if update_koima_index is not None:  # 순수 추가: KOIMA 부문별 지수도 업데이트에 덧붙임
    FETCHERS["koima_index"] = update_koima_index


# ── 수집 실행: 서버별 그룹 병렬 + 월간 데이터 당일 캐시 ────────────────────
# 같은 외부 서버를 치는 수집기는 한 그룹에 묶어 그룹 '안에서는 순차'로 돌린다.
# 서로 다른 그룹끼리만 동시에 실행한다.
FETCH_GROUPS = [
    ["simmons_news", "domestic", "global_brands"],   # news.google.com
    ["schedule_reliability", "sr_forecast"],         # sea-intelligence.com
    ["oil_prices", "oil_forecast"],                  # petronet.co.kr
    ["usd_krw", "fx"],                               # frankfurter.app
    ["koima_index"],                                 # koimaindex.com
    ["competitors"],                                 # SEC EDGAR
    ["icis_forecast"],                               # 네트워크 없음(고정 데이터)
]
# 한 달에 한 번만 바뀌는 데이터 → 같은 날 이미 받았으면 재사용한다.
MONTHLY_CACHED = ("koima_index", "schedule_reliability", "sr_forecast")
_UPDATE_CACHE = "update-cache.json"


def _cache_path():
    """쓰기 가능한 위치를 고른다(배포 번들은 읽기 전용이라 임시 디렉터리로 폴백)."""
    base = os.path.dirname(os.path.abspath(__file__))
    d = os.path.join(base, "tmp")
    try:
        if not os.path.isdir(d):
            os.makedirs(d)
        return os.path.join(d, _UPDATE_CACHE)
    except OSError:
        import tempfile
        return os.path.join(tempfile.gettempdir(), _UPDATE_CACHE)


def _cache_load():
    """{name: {"date": "YYYY-MM-DD", "payload": {...}}}. 실패해도 예외를 내지 않는다."""
    try:
        p = _cache_path()
        if os.path.isfile(p):
            return json.load(io.open(p, encoding="utf-8"))
    except Exception:  # noqa: BLE001
        pass
    return {}


def _cache_save(cache):
    try:
        with io.open(_cache_path(), "w", encoding="utf-8") as f:
            f.write(json.dumps(cache, ensure_ascii=False))
    except Exception as e:  # noqa: BLE001
        print("[update] 캐시 저장 생략(쓰기 불가): %s" % e)


def run_fetchers(force=False):
    """FETCHERS 를 그룹 병렬로 실행해 sections dict 반환.
       force=True 면 월간 캐시를 무시하고 전부 새로 수집한다."""
    groups = [[n for n in g if n in FETCHERS] for g in FETCH_GROUPS]
    listed = {n for g in groups for n in g}
    groups += [[n] for n in FETCHERS if n not in listed]   # 목록에 없는 수집기는 단독 그룹
    groups = [g for g in groups if g]

    today = datetime.date.today().isoformat()
    cache = {} if force else _cache_load()
    sections, used_cache = {}, []

    def run_group(names):
        out = {}
        for n in names:
            c = cache.get(n)
            if (not force) and n in MONTHLY_CACHED and c and c.get("date") == today                     and isinstance(c.get("payload"), dict):
                out[n] = c["payload"]
                used_cache.append(n)
                continue
            t = time.time()
            try:
                out[n] = FETCHERS[n]()
            except Exception as e:  # noqa: BLE001
                out[n] = {"status": "error", "reason": str(e)}
            st = out[n].get("status", "?") if isinstance(out[n], dict) else "?"
            print("[update] %-22s %6.1f초  %s" % (n, time.time() - t, st))
        return out

    t0 = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, len(groups))) as ex:
        for res in ex.map(run_group, groups):
            sections.update(res)

    if used_cache:
        print("[update] 당일 캐시 재사용(월간 데이터): %s" % ", ".join(used_cache))
    new_cache = dict(cache)
    for n in MONTHLY_CACHED:
        s = sections.get(n)
        if isinstance(s, dict) and s.get("status") == "ok" and n not in used_cache:
            new_cache[n] = {"date": today, "payload": s}
    if new_cache != cache:
        _cache_save(new_cache)

    failed = [n for n, s in sections.items()
              if isinstance(s, dict) and s.get("status") not in ("ok", None)]
    print("[update] 총 %.1f초 · 그룹 %d개 병렬 · 실패 %s"
          % (time.time() - t0, len(groups), ", ".join(failed) if failed else "없음"))
    return sections


def build_payload(force=False):
    """모든 fetcher 실행 → dashboard_server.py 와 동일한 응답 dict."""
    sections = run_fetchers(force=force)
    updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {"updated_at": updated_at, "sections": sections}


# ── Vercel 서버리스 진입점 ────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def _respond(self):
        try:
            payload = build_payload()
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
        except Exception as e:  # noqa: BLE001 — 최후 방어: 절대 500 raw 로 죽지 않게
            body = json.dumps({"error": str(e)}, ensure_ascii=False).encode("utf-8")
            self.send_response(500)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", "*")  # 다른 출처에서 열어도 허용
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        self._respond()

    def do_POST(self):
        self._respond()
